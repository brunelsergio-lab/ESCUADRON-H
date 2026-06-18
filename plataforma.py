# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from zoneinfo import ZoneInfo
import os
from io import BytesIO
import base64
import json
import hmac
import re
import secrets as py_secrets
import unicodedata
from html import escape
import streamlit.components.v1 as components

LIVE_SEARCH_COMPONENT = components.declare_component(
    "live_search_input",
    path=os.path.join(os.path.dirname(__file__), "live_search_component")
)

def live_search_input(label, placeholder, key):
    value = LIVE_SEARCH_COMPONENT(
        label=label,
        placeholder=placeholder,
        value=st.session_state.get(key, ""),
        key=f"{key}_live",
        default=st.session_state.get(key, "")
    )
    value = value or ""
    st.session_state[key] = value
    return value

APP_LOGO_FILE = "logo nuevo.jpg"
APP_TZ = ZoneInfo("America/Argentina/Buenos_Aires")

def ahora_local():
    return datetime.now(APP_TZ)


def excel_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

EXCEL_FONT_NAME = "Arial"
EXCEL_OLIVE_DARK = "4B5320"
EXCEL_OLIVE = "556B2F"
EXCEL_OLIVE_LIGHT = "E8EAD7"
EXCEL_OLIVE_ROW = "F4F6ED"
EXCEL_TEXT_DARK = "1F2937"
EXCEL_TEXT_MUTED = "555555"
EXCEL_BORDER = "A6A078"
EXCEL_WHITE = "FFFFFF"


def excel_font(**kwargs):
    kwargs.setdefault("name", EXCEL_FONT_NAME)
    return Font(**kwargs)

def descargar_archivo_auto(data, file_name, mime):
    href = f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"
    components.html(
        f"""
        <script>
        const a = document.createElement('a');
        a.href = {json.dumps(href)};
        a.download = {json.dumps(file_name)};
        document.body.appendChild(a);
        a.click();
        a.remove();
        </script>
        """,
        height=0,
    )

DIAS_SEMANA = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}

ESTADOS_AUSENCIA = {"AUSENTE", "ART", "DAF", "LES", "SSD", "LAO", "AUTORIZADO", "DESCANSO DE GUARDIA"}
ESTADOS_GUARDIA_DIURNA = {"ENTRANTE GUARDIA DIURNA", "SERVICIO DE ARMAS DIURNA"}
ESTADOS_GUARDIA_NOCTURNA = {"ENTRANTE GUARDIA NOCTURNA", "SERVICIO DE ARMAS NOCTURNA"}
ESTADOS_COMISION = {"COMISION", "COMISION DE SERVICIO"}
AMBITOS_NOVEDAD = {
    "AUSENTE": "Ausente",
    "INSTITUTO": "Presente en instituto",
    "ESCUADRON": "Presente en escuadrón",
}

def ambito_por_defecto(estado):
    if estado in ESTADOS_AUSENCIA:
        return "AUSENTE"
    if estado == "PRESENTE EN INSTITUTO":
        return "INSTITUTO"
    if estado == "PRESENTE EN ESCUADRÓN":
        return "ESCUADRON"
    if estado == "COMISIÓN":
        return "INSTITUTO"
    return "ESCUADRON"

def ambito_efectivo(novedad):
    estado = novedad.get('estado', '')
    ambito_guardado = novedad.get('ambito')
    ambito_estado = ambito_por_defecto(estado)
    if estado in {"AUSENTE", "PRESENTE EN INSTITUTO", "PRESENTE EN ESCUADRÓN"}:
        return ambito_estado
    return ambito_guardado or ambito_estado

def es_tercer_anio(valor):
    return "III" in str(valor).upper() or "TERCER" in str(valor).upper()

def es_aop(valor):
    texto = str(valor).upper()
    return "AOP" in texto or "CAO" in texto or "AUXILIAR" in texto


def normalizar_hora_ingreso(valor):
    texto = str(valor or "06:00").strip().replace(".", ":")
    if not texto:
        return "06:00"
    digitos = "".join(ch for ch in texto if ch.isdigit())
    if ":" not in texto and len(digitos) in {3, 4}:
        digitos = digitos.zfill(4)
        return f"{digitos[:2]}:{digitos[2:]}"
    if ":" in texto:
        partes = texto.split(":", 1)
        if all(p.strip().isdigit() for p in partes):
            return f"{int(partes[0]):02d}:{int(partes[1]):02d}"
    return texto


def hora_ingreso_aula(aula):
    horarios = st.session_state.get("horarios_config", {})
    cfg = horarios.get(aula) or horarios.get(normalizar_aula(aula)) or {}
    return normalizar_hora_ingreso(cfg.get("ent_m", "06:00"))


def aula_ingresa_primera_obligacion(aula, ahora=None, fecha_objetivo=None):
    return hora_ingreso_aula(aula) == "06:00"


def hora_a_minutos(valor, defecto="00:00"):
    hora = normalizar_hora_ingreso(valor or defecto)
    try:
        hh, mm = hora.split(":", 1)
        return int(hh) * 60 + int(mm)
    except Exception:
        hh, mm = defecto.split(":", 1)
        return int(hh) * 60 + int(mm)


def minuto_operativo_reporte():
    fecha = st.session_state.get("fecha_reporte", ahora_local().date())
    hoy = ahora_local().date()
    if fecha < hoy:
        return 23 * 60 + 59
    if fecha > hoy:
        return 0
    ahora = ahora_local()
    return ahora.hour * 60 + ahora.minute


def estado_programado_aula(aula, prefijo, cfg_estado=None):
    horarios = st.session_state.get("horarios_config", {})
    hor = horarios.get(aula) or horarios.get(normalizar_aula(aula)) or {}
    ent = hora_a_minutos(hor.get(f"ent_{prefijo}", "06:00"), "06:00" if prefijo == "m" else "13:00")
    sal = hora_a_minutos(hor.get(f"sal_{prefijo}", "12:00"), "12:00" if prefijo == "m" else "19:00")
    actual = minuto_operativo_reporte()
    if actual < ent or actual >= sal:
        return "FUERA", "FRANCO"
    if prefijo == "m" and normalizar_hora_ingreso(hor.get("ent_m", "06:00")) == "06:00" and actual < 7 * 60:
        return "EN INSTITUTO", "FORMACION"
    ubic_key = f"ubicacion_{prefijo}"
    previa = (cfg_estado or {}).get(ubic_key, "EN AULA")
    if previa in {"FRANCO", "FORMACION"}:
        previa = "EN AULA"
    return "EN INSTITUTO", previa or "EN AULA"


def sincronizar_ubicacion_con_horarios():
    for aula, cfg in st.session_state.get("estado_aulas", {}).items():
        for prefijo in ("m", "t"):
            estado_key = f"estado_{prefijo}"
            ubic_key = f"ubicacion_{prefijo}"
            salida_key = f"salida_{prefijo}"
            estado_auto, ubic_auto = estado_programado_aula(aula, prefijo, cfg)
            cfg[estado_key] = estado_auto
            cfg[ubic_key] = ubic_auto
            if estado_auto == "FUERA":
                cfg[salida_key] = "FRANCO"
            elif cfg.get(salida_key) == "FRANCO":
                cfg[salida_key] = None

def numero_letras(n):
    mapa = {
        0: "CERO", 1: "UN", 2: "DOS", 3: "TRES", 4: "CUATRO", 5: "CINCO",
        6: "SEIS", 7: "SIETE", 8: "OCHO", 9: "NUEVE", 10: "DIEZ",
        11: "ONCE", 12: "DOCE", 13: "TRECE", 14: "CATORCE", 15: "QUINCE",
        16: "DIECISÉIS", 17: "DIECISIETE", 18: "DIECIOCHO", 19: "DIECINUEVE",
        20: "VEINTE"
    }
    return mapa.get(int(n), str(n))

def normalizar_estado_novedad(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().upper())
    return "".join(ch for ch in texto if not unicodedata.combining(ch))


def formatear_lista_novedades(novedades, estado, curso_fn):
    filtradas = [n for n in novedades if n['estado'] == estado and curso_fn(n.get('grado', ''))]
    if not filtradas:
        return ".-"
    lineas = []
    for idx, nov in enumerate(filtradas, 1):
        grado = "ASP III ANO" if es_tercer_anio(nov.get('grado', '')) else "ASP I"
        detalle = f" {nov.get('detalle', '').strip()}" if nov.get('detalle') else ""
        lineas.append(f"{idx}. {grado} {nov['nombre']}{detalle} (D: {nov['fecha_ini']}, H: {nov['fecha_fin']})")
    return "\n".join(lineas)

def formatear_servicio(novedades, estados, curso_fn, titulo):
    estados_norm = {normalizar_estado_novedad(e) for e in estados}
    filtradas = [
        n for n in novedades
        if normalizar_estado_novedad(n.get('estado')) in estados_norm and curso_fn(n.get('grado', ''))
    ]
    if not filtradas:
        return f"▫️ {titulo}:"
    plural = "ASPIRANTES" if len(filtradas) != 1 else "ASPIRANTE"
    lineas = [f"▫️ {titulo}: {numero_letras(len(filtradas))} ({len(filtradas)}) {plural}."]
    for idx, nov in enumerate(filtradas, 1):
        grado = "ASP III ANO" if es_tercer_anio(nov.get('grado', '')) else "ASP I"
        detalle = f" {nov.get('detalle', '').strip()}" if nov.get('detalle') else ""
        lineas.append(f"{idx}. {grado} {nov['nombre']}{detalle}")
    return "\n".join(lineas)


def formatear_ingreso_diferenciado(df_presentes_escuadron, curso_fn):
    lineas = []
    for aula in AULAS_UNICAS:
        cfg = st.session_state.get("horarios_config", {}).get(aula, {})
        hora = normalizar_hora_ingreso(cfg.get("ent_m", "06:00"))
        if hora == "06:00":
            continue
        alumnos = df_presentes_escuadron[
            (df_presentes_escuadron["AULA"] == aula) &
            (df_presentes_escuadron["GRADO"].map(curso_fn))
        ]
        cant = len(alumnos)
        if cant:
            lineas.append(f"INGRESO {hora} HS: {numero_letras(cant)} ({cant}) aspirante(s) del aula {aula}.")
    return "\n".join(lineas) if lineas else ".-"


def generar_minuta_formal_clasica():
    fecha_minuta = st.session_state.fecha_reporte.strftime('%d%b%y').upper()

    # La minuta debe reflejar la base actual, no el estado que quedo en pantalla.
    novedades = obtener_novedades(FECHA_PARTE_STR)
    estado_asistencia_actual = obtener_asistencia(FECHA_PARTE_STR)
    st.session_state.novedades_lista = novedades
    st.session_state.estado_asistencia = estado_asistencia_actual

    ambito_minuta = {
        n['orden']: ambito_efectivo(n)
        for n in novedades
    }
    ausentes_novedad = {orden for orden, ambito in ambito_minuta.items() if ambito == "AUSENTE"}
    presentes_instituto_novedad = {orden for orden, ambito in ambito_minuta.items() if ambito in {"INSTITUTO", "ESCUADRON"}}
    presentes_escuadron_novedad = {orden for orden, ambito in ambito_minuta.items() if ambito == "ESCUADRON"}
    ausentes_manuales = {orden for orden, estado in estado_asistencia_actual.items() if estado == "AUSENTE"}
    presentes_instituto_manuales = {
        orden for orden, estado in estado_asistencia_actual.items()
        if estado in {"PRESENTE", "PRESENTE EN INSTITUTO", "PRESENTE EN ESCUADRÓN"}
    }
    presentes_escuadron_manuales = {
        orden for orden, estado in estado_asistencia_actual.items()
        if estado in {"PRESENTE", "PRESENTE EN ESCUADRÓN"}
    }

    total_ausentes_minuta = ausentes_novedad | (ausentes_manuales - presentes_instituto_manuales)

    df_presentes_instituto_minuta = df[
        ~df['ORDEN_LIMP'].isin(total_ausentes_minuta)
    ]
    df_presentes_escuadron_minuta = df[
        (~df['ORDEN_LIMP'].isin(total_ausentes_minuta)) &
        (
            df['ORDEN_LIMP'].isin(presentes_escuadron_novedad) |
            df['ORDEN_LIMP'].isin(presentes_escuadron_manuales) |
            ((~df['ORDEN_LIMP'].isin(ambito_minuta.keys())) &
             (~df['ORDEN_LIMP'].isin(presentes_instituto_manuales)))
        )
    ]
    df_presentes_primera_minuta = df_presentes_escuadron_minuta[
        df_presentes_escuadron_minuta['AULA'].map(aula_ingresa_primera_obligacion)
    ]

    df_tercero = df[df['GRADO'].map(es_tercer_anio)]
    df_aop = df[df['GRADO'].map(es_aop)]
    ausentes_tercero = set(df_tercero['ORDEN_LIMP']) & total_ausentes_minuta
    ausentes_aop = set(df_aop['ORDEN_LIMP']) & total_ausentes_minuta
    presentes_instituto_tercero = len(df_presentes_instituto_minuta[df_presentes_instituto_minuta['GRADO'].map(es_tercer_anio)])
    presentes_instituto_aop = len(df_presentes_instituto_minuta[df_presentes_instituto_minuta['GRADO'].map(es_aop)])
    presentes_escuadron_tercero = len(df_presentes_escuadron_minuta[df_presentes_escuadron_minuta['GRADO'].map(es_tercer_anio)])
    presentes_escuadron_aop = len(df_presentes_escuadron_minuta[df_presentes_escuadron_minuta['GRADO'].map(es_aop)])
    formados_tercero = len(df_presentes_primera_minuta[df_presentes_primera_minuta['GRADO'].map(es_tercer_anio)])
    formados_aop = len(df_presentes_primera_minuta[df_presentes_primera_minuta['GRADO'].map(es_aop)])
    presentes_instituto_total_minuta = len(df_presentes_instituto_minuta)
    presentes_escuadron_total_minuta = len(df_presentes_escuadron_minuta)
    primera_total_minuta = len(df_presentes_primera_minuta)

    lineas = [
        f'MINUTA INFORMATIVA DEL ESCUADRÓN H "CABO MARCELO GODOY" DEL DÍA {fecha_minuta}',
        "",
        f"FE: {TOTAL_ESCUADRON}",
        f"PRESENTES EN INSTITUTO: {presentes_instituto_total_minuta}",
        f"PRESENTES EN ESCUADRON: {presentes_escuadron_total_minuta}",
        f"AUSENTES / FUERA DEL INSTITUTO: {len(total_ausentes_minuta)}",
        f"FORMADOS A PRIMERA OBLIGACIÓN: {primera_total_minuta}",
        "",
        "✅ CURSO DE TERCER AÑO",
        "",
        f"FE: {len(df_tercero)}",
        f"PRESENTES EN INSTITUTO: {presentes_instituto_tercero}",
        f"PRESENTES EN ESCUADRON: {presentes_escuadron_tercero}",
        f"AUSENTES / FUERA DEL INSTITUTO: {len(ausentes_tercero)}",
        f"FORMADOS PRIMERA OBLIGACIÓN: {formados_tercero}",
        "",
        "OBS:",
        "",
        "▫️ INGRESO HORARIO DIFERENCIADO:",
        formatear_ingreso_diferenciado(df_presentes_escuadron_minuta, es_tercer_anio),
        formatear_servicio(novedades, ESTADOS_COMISION, es_tercer_anio, "COMISION DE SERVICIO"),
        "",
        formatear_servicio(novedades, ESTADOS_GUARDIA_DIURNA, es_tercer_anio, "SERVICIO DE ARMAS DIURNA"),
        "",
        formatear_servicio(novedades, ESTADOS_GUARDIA_NOCTURNA | {"DESCANSO DE GUARDIA"}, es_tercer_anio, "DESCANSO DE SERVICIO DE ARMAS NOCTURNO"),
        "",
        "NOVEDADES SANITARIAS:",
        "",
        "▫️ SIN SERVICIO EN DOMICILIO:",
        formatear_lista_novedades(novedades, "SSD", es_tercer_anio),
        "",
        "▫️ ART:",
        formatear_lista_novedades(novedades, "ART", es_tercer_anio),
        "",
        "▫️ DAF:",
        formatear_lista_novedades(novedades, "DAF", es_tercer_anio),
        "",
        "▫️ AUTORIZADO:",
        formatear_lista_novedades(novedades, "AUTORIZADO", es_tercer_anio),
        "",
        "✅ CURSO AUXILIAR OPERATIVO",
        f"FE: {len(df_aop)}",
        f"PRESENTES EN INSTITUTO: {presentes_instituto_aop}",
        f"PRESENTES EN ESCUADRON: {presentes_escuadron_aop}",
        f"AUSENTES / FUERA DEL INSTITUTO: {len(ausentes_aop)}",
        f"FORMADOS PRIMERA OBLIGACIÓN: {formados_aop}",
        "",
        "OBS:",
        "",
        "▫️ INGRESO EN HORARIO DIFERENCIAL:",
        formatear_ingreso_diferenciado(df_presentes_escuadron_minuta, es_aop),
        formatear_servicio(novedades, ESTADOS_COMISION, es_aop, "COMISION DE SERVICIO"),
        "",
        formatear_servicio(novedades, ESTADOS_GUARDIA_DIURNA, es_aop, "SERVICIO DE ARMAS DIURNO"),
        "",
        formatear_servicio(novedades, ESTADOS_GUARDIA_NOCTURNA | {"DESCANSO DE GUARDIA"}, es_aop, "DESCANSO DE SERVICIO DE ARMAS NOCTURNO"),
        "",
        "▫️ ART",
        formatear_lista_novedades(novedades, "ART", es_aop),
        "",
        "▫️ SIN SERVICIO EN DOMICILIO:",
        formatear_lista_novedades(novedades, "SSD", es_aop),
        "",
        "▫️ LAO: (A CUENTA DE LAO)",
        formatear_lista_novedades(novedades, "LAO", es_aop),
        "",
        "▫️ LES:",
        formatear_lista_novedades(novedades, "LES", es_aop),
        "",
        "▫️ DAF:",
        formatear_lista_novedades(novedades, "DAF", es_aop),
    ]
    return "\n".join(lineas)



def separador_minuta():
    return "\u2501" * 22


def grado_minuta(nov):
    return "ASP III A\u00d1O" if es_tercer_anio(nov.get('grado', '')) else "ASP I"


def filtrar_novedades_minuta(novedades, estados, curso_fn):
    estados_norm = {normalizar_estado_novedad(e) for e in estados}
    return [
        n for n in novedades
        if normalizar_estado_novedad(n.get('estado')) in estados_norm and curso_fn(n.get('grado', ''))
    ]


def bloque_resumen_general_minuta(total, presentes_instituto, presentes_escuadron, ausentes, formados):
    return "\n".join([
        "\U0001F530 RESUMEN GENERAL",
        f"\U0001F465 FE: {total}",
        f"\U0001F3EB Presentes en Instituto: {presentes_instituto}",
        f"\U0001F7E2 Presentes en Escuadr\u00f3n: {presentes_escuadron}",
        f"\U0001F534 Ausentes / Fuera del Instituto: {ausentes}",
        f"\U0001F9CD Formados a 1ra Obligaci\u00f3n: {formados}",
    ])


def lineas_ingresos_diferenciados_minuta(df_presentes_escuadron, curso_fn):
    lineas = []
    for aula in AULAS_UNICAS:
        cfg = st.session_state.get("horarios_config", {}).get(aula, {})
        hora = normalizar_hora_ingreso(cfg.get("ent_m", "06:00"))
        if hora == "06:00":
            continue
        alumnos = df_presentes_escuadron[
            (df_presentes_escuadron["AULA"] == aula) &
            (df_presentes_escuadron["GRADO"].map(curso_fn))
        ]
        cant = len(alumnos)
        if cant:
            lineas.append(f"\u25ab\ufe0f Aula {aula}: {cant} aspirante(s) \u2014 {hora} hs")
    return lineas


def bloque_lista_operativa_minuta(titulo, icono, novedades, estados, curso_fn, etiqueta_detalle="Detalle"):
    filtradas = filtrar_novedades_minuta(novedades, estados, curso_fn)
    if not filtradas:
        return ""
    lineas = [f"{icono} {titulo}", f"Total: {len(filtradas)} aspirante(s)", ""]
    for idx, nov in enumerate(filtradas, 1):
        detalle = str(nov.get('detalle') or '').strip()
        fechas = f"Desde: {nov.get('fecha_ini', '-')} | Hasta: {nov.get('fecha_fin', '-')}"
        lineas.append(f"{idx}. {grado_minuta(nov)} {nov['nombre']}")
        if detalle:
            lineas.append(f"   {etiqueta_detalle}: {detalle} | {fechas}")
        else:
            lineas.append(f"   {fechas}")
    return "\n".join(lineas)


def bloque_novedades_operativas_minuta(novedades, df_presentes_escuadron, curso_fn):
    bloques = ["\U0001F4DD NOVEDADES OPERATIVAS"]

    ingresos = lineas_ingresos_diferenciados_minuta(df_presentes_escuadron, curso_fn)
    if ingresos:
        bloques.append("\n".join(["\U0001F558 INGRESOS DIFERENCIADOS", *ingresos]))

    for bloque in [
        bloque_lista_operativa_minuta("COMISI\u00d3N DE SERVICIO", "\U0001F396\ufe0f", novedades, ESTADOS_COMISION, curso_fn, "Comisi\u00f3n"),
        bloque_lista_operativa_minuta("SERVICIO DE ARMAS DIURNO", "\U0001F6E1\ufe0f", novedades, ESTADOS_GUARDIA_DIURNA, curso_fn),
        bloque_lista_operativa_minuta("DESCANSO DE SERVICIO DE ARMAS NOCTURNO", "\U0001F319", novedades, ESTADOS_GUARDIA_NOCTURNA | {"DESCANSO DE GUARDIA"}, curso_fn),
    ]:
        if bloque:
            bloques.append(bloque)

    if len(bloques) == 1:
        bloques.append("Sin novedades operativas.")
    return "\n\n".join(bloques)


def bloque_sanitario_categoria_minuta(titulo, novedades, estado, curso_fn):
    filtradas = filtrar_novedades_minuta(novedades, {estado}, curso_fn)
    if not filtradas:
        return ""
    lineas = [f"\u25ab\ufe0f {titulo}", ""]
    for idx, nov in enumerate(filtradas, 1):
        detalle = str(nov.get('detalle') or '').strip()
        fechas = f"Desde: {nov.get('fecha_ini', '-')} | Hasta: {nov.get('fecha_fin', '-')}"
        lineas.append(f"{idx}. {grado_minuta(nov)} {nov['nombre']}")
        if detalle:
            lineas.append(f"   {detalle} | {fechas}")
        else:
            lineas.append(f"   {fechas}")
    return "\n".join(lineas)


def bloque_novedades_sanitarias_minuta(novedades, curso_fn):
    bloques = []
    categorias = [
        ("SIN SERVICIO EN DOMICILIO", "SSD"),
        ("ART", "ART"),
        ("DAF", "DAF"),
        ("LAO", "LAO"),
        ("LES", "LES"),
        ("AUTORIZADO", "AUTORIZADO"),
    ]
    for titulo, estado in categorias:
        bloque = bloque_sanitario_categoria_minuta(titulo, novedades, estado, curso_fn)
        if bloque:
            bloques.append(bloque)
    if not bloques:
        return "\U0001F3E5 NOVEDADES SANITARIAS\nSin otras novedades sanitarias."
    return "\U0001F3E5 NOVEDADES SANITARIAS\n\n" + "\n\n".join(bloques)


def bloque_curso_minuta(titulo, fe, presentes_instituto, presentes_escuadron, ausentes, formados, novedades, df_presentes_escuadron, curso_fn):
    return "\n".join([
        separador_minuta(),
        f"\u2705 {titulo}",
        separador_minuta(),
        "",
        "\U0001F4CA PRESENTISMO:",
        f"\U0001F465 FE: {fe}",
        f"\U0001F3EB Presentes Instituto: {presentes_instituto}",
        f"\U0001F7E2 Presentes Escuadr\u00f3n: {presentes_escuadron}",
        f"\U0001F534 Ausentes: {ausentes}",
        f"\U0001F9CD Formados 1ra Obligaci\u00f3n: {formados}",
        "",
        bloque_novedades_operativas_minuta(novedades, df_presentes_escuadron, curso_fn),
        "",
        bloque_novedades_sanitarias_minuta(novedades, curso_fn),
    ])


def generar_minuta_visual_whatsapp():
    fecha_minuta = st.session_state.fecha_reporte.strftime('%d%b%y').upper()

    novedades = obtener_novedades(FECHA_PARTE_STR)
    estado_asistencia_actual = obtener_asistencia(FECHA_PARTE_STR)
    st.session_state.novedades_lista = novedades
    st.session_state.estado_asistencia = estado_asistencia_actual

    ambito_minuta = {n['orden']: ambito_efectivo(n) for n in novedades}
    ausentes_novedad = {orden for orden, ambito in ambito_minuta.items() if ambito == "AUSENTE"}
    presentes_instituto_manuales = {
        orden for orden, estado in estado_asistencia_actual.items()
        if estado in {"PRESENTE", "PRESENTE EN INSTITUTO", "PRESENTE EN ESCUADR?N"}
    }
    presentes_escuadron_novedad = {orden for orden, ambito in ambito_minuta.items() if ambito == "ESCUADRON"}
    presentes_escuadron_manuales = {
        orden for orden, estado in estado_asistencia_actual.items()
        if estado in {"PRESENTE", "PRESENTE EN ESCUADR?N"}
    }
    ausentes_manuales = {orden for orden, estado in estado_asistencia_actual.items() if estado == "AUSENTE"}
    total_ausentes_minuta = ausentes_novedad | (ausentes_manuales - presentes_instituto_manuales)

    df_presentes_instituto_minuta = df[~df['ORDEN_LIMP'].isin(total_ausentes_minuta)]
    df_presentes_escuadron_minuta = df[
        (~df['ORDEN_LIMP'].isin(total_ausentes_minuta)) &
        (
            df['ORDEN_LIMP'].isin(presentes_escuadron_novedad) |
            df['ORDEN_LIMP'].isin(presentes_escuadron_manuales) |
            ((~df['ORDEN_LIMP'].isin(ambito_minuta.keys())) &
             (~df['ORDEN_LIMP'].isin(presentes_instituto_manuales)))
        )
    ]
    df_presentes_primera_minuta = df_presentes_escuadron_minuta[
        df_presentes_escuadron_minuta['AULA'].map(aula_ingresa_primera_obligacion)
    ]

    df_tercero = df[df['GRADO'].map(es_tercer_anio)]
    df_aop = df[df['GRADO'].map(es_aop)]
    ausentes_tercero = set(df_tercero['ORDEN_LIMP']) & total_ausentes_minuta
    ausentes_aop = set(df_aop['ORDEN_LIMP']) & total_ausentes_minuta
    presentes_instituto_tercero = len(df_presentes_instituto_minuta[df_presentes_instituto_minuta['GRADO'].map(es_tercer_anio)])
    presentes_instituto_aop = len(df_presentes_instituto_minuta[df_presentes_instituto_minuta['GRADO'].map(es_aop)])
    presentes_escuadron_tercero = len(df_presentes_escuadron_minuta[df_presentes_escuadron_minuta['GRADO'].map(es_tercer_anio)])
    presentes_escuadron_aop = len(df_presentes_escuadron_minuta[df_presentes_escuadron_minuta['GRADO'].map(es_aop)])
    formados_tercero = len(df_presentes_primera_minuta[df_presentes_primera_minuta['GRADO'].map(es_tercer_anio)])
    formados_aop = len(df_presentes_primera_minuta[df_presentes_primera_minuta['GRADO'].map(es_aop)])

    bloques = [
        "\U0001F4CC MINUTA INFORMATIVA",
        "ESCUADR\u00d3N H \u201cCABO MARCELO GODOY\u201d",
        f"\U0001F4C5 Fecha: {fecha_minuta}",
        "",
        bloque_resumen_general_minuta(
            TOTAL_ESCUADRON,
            len(df_presentes_instituto_minuta),
            len(df_presentes_escuadron_minuta),
            len(total_ausentes_minuta),
            len(df_presentes_primera_minuta),
        ),
        "",
        bloque_curso_minuta(
            "CURSO DE TERCER A\u00d1O",
            len(df_tercero),
            presentes_instituto_tercero,
            presentes_escuadron_tercero,
            len(ausentes_tercero),
            formados_tercero,
            novedades,
            df_presentes_escuadron_minuta,
            es_tercer_anio,
        ),
        "",
        bloque_curso_minuta(
            "CURSO AUXILIAR OPERATIVO",
            len(df_aop),
            presentes_instituto_aop,
            presentes_escuadron_aop,
            len(ausentes_aop),
            formados_aop,
            novedades,
            df_presentes_escuadron_minuta,
            es_aop,
        ),
    ]
    return "\n".join(bloques)


def generar_minuta_informativa(formato="Visual WhatsApp / Celular"):
    if formato == "Formal cl\u00e1sico":
        return generar_minuta_formal_clasica()
    return generar_minuta_visual_whatsapp()

def normalizar_aula(aula):
    return str(aula).strip().upper().replace(" ", "")

def cargar_horarios_txt(path):
    if not os.path.exists(path):
        return 0

    df_horarios = pd.read_csv(path, encoding="utf-8-sig")
    df_horarios.columns = df_horarios.columns.str.strip()
    df_horarios = df_horarios.rename(columns={
        "Aula": "aula",
        "Dia": "dia",
        "Entrada_Mañana": "ent_m",
        "Salida_Mañana": "sal_m",
        "Entrada_Tarde": "ent_t",
        "Salida_Tarde": "sal_t",
    })

    required = {"aula", "dia", "ent_m", "sal_m", "ent_t", "sal_t"}
    if not required.issubset(set(df_horarios.columns)):
        return 0

    total = 0
    for _, row in df_horarios.iterrows():
        guardar_horarios_dia(normalizar_aula(row["aula"]), str(row["dia"]).strip(), {
            "ent_m": str(row["ent_m"]).strip(),
            "sal_m": str(row["sal_m"]).strip(),
            "ent_t": str(row["ent_t"]).strip(),
            "sal_t": str(row["sal_t"]).strip(),
        })
        total += 1
    return total

# 🔹 IMPORTS CORREGIDOS (NO OMITIR NINGUNA FUNCIÓN)
import db_manager as _db_manager
from db_manager import (
    init_db,
    obtener_novedades, agregar_novedad, actualizar_novedad, eliminar_novedad, vaciar_novedades,
    obtener_estado_aulas, guardar_estado_aula,
    obtener_almuerzo, agregar_almuerzo, quitar_almuerzo,
    obtener_horarios, guardar_horarios, obtener_horarios_dia, guardar_horarios_dia,
    obtener_config, guardar_config,
    obtener_asistencia, actualizar_asistencia,
    obtener_contacto, obtener_todos_contactos, guardar_contacto,
    registrar_movimiento, obtener_movimientos
)
st.set_page_config(
    page_title="Gestión de Parte Diario - Escuadrón H",
    page_icon="🟡",
    layout="wide",
    initial_sidebar_state="expanded"
)


def load_css(file_path):
    """Carga estilos visuales desde un archivo CSS externo."""
    try:
        with open(file_path, "r", encoding="utf-8") as css_file:
            st.markdown(f"<style>{css_file.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        pass


load_css(os.path.join(os.path.dirname(__file__), "assets", "styles.css"))


def activar_modo_app_movil():
    """Agrega configuración visual y PWA básica para instalar en teléfono."""
    st.markdown(
        """
        <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
        <meta name="theme-color" content="#F1B82D">
        <meta name="apple-mobile-web-app-capable" content="yes">
        <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
        <meta name="apple-mobile-web-app-title" content="Escuadrón H">
        <link rel="apple-touch-icon" href="/app/static/icon-192.png">
        <link rel="manifest" href="/app/static/manifest.json">
        <style>
        @media (max-width: 768px) {
            .block-container {
                padding: 0.65rem 0.55rem 5rem 0.55rem !important;
                max-width: 100% !important;
            }
            h1, h2, h3 { line-height: 1.15 !important; }
            h1 { font-size: 1.45rem !important; }
            h2 { font-size: 1.25rem !important; }
            h3 { font-size: 1.08rem !important; }
            [data-testid="stHorizontalBlock"] {
                gap: 0.35rem !important;
            }
            [data-testid="column"] {
                min-width: min(100%, 360px) !important;
            }
            .stButton > button, .stDownloadButton > button {
                min-height: 42px !important;
                border-radius: 12px !important;
                font-size: 0.92rem !important;
            }
            .stTabs [data-baseweb="tab-list"] {
                overflow-x: auto !important;
                white-space: nowrap !important;
                gap: 0.25rem !important;
            }
            .stTabs [data-baseweb="tab"] {
                padding: 0.55rem 0.7rem !important;
                min-width: max-content !important;
            }
            [data-testid="stDataFrame"], [data-testid="stTable"] {
                font-size: 0.78rem !important;
            }
            .stTextInput input, .stSelectbox div, .stDateInput input, textarea {
                font-size: 16px !important;
            }
            [data-testid="stSidebar"] {
                width: min(88vw, 330px) !important;
            }
            .login-title { font-size: 1.65rem !important; }
            .login-subtitle { font-size: 0.95rem !important; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    components.html(
        """
        <script>
        (function() {
            const doc = window.parent.document;
            function ensureLink(rel, href) {
                if (!doc.querySelector(`link[rel="${rel}"][href="${href}"]`)) {
                    const link = doc.createElement('link');
                    link.rel = rel;
                    link.href = href;
                    doc.head.appendChild(link);
                }
            }
            ensureLink('manifest', '/app/static/manifest.json');
            ensureLink('apple-touch-icon', '/app/static/icon-192.png');
            let metaTheme = doc.querySelector('meta[name="theme-color"]');
            if (!metaTheme) {
                metaTheme = doc.createElement('meta');
                metaTheme.name = 'theme-color';
                doc.head.appendChild(metaTheme);
            }
            metaTheme.content = '#F1B82D';
            if ('serviceWorker' in window.parent.navigator) {
                window.parent.navigator.serviceWorker.register('/app/static/service-worker.js').catch(() => {});
            }
        })();
        </script>
        """,
        height=0,
    )

activar_modo_app_movil()



def asset_data_uri(relative_path):
    """Convierte un archivo local peque?o en data URI para usarlo dentro de HTML."""
    try:
        full_path = os.path.join(os.path.dirname(__file__), relative_path)
        ext = os.path.splitext(relative_path)[1].lower()
        mime = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
        }.get(ext, "application/octet-stream")
        with open(full_path, "rb") as file:
            encoded = base64.b64encode(file.read()).decode("ascii")
        return f"data:{mime};base64,{encoded}"
    except Exception:
        return ""

actualizar_movimiento = getattr(_db_manager, "actualizar_movimiento", None)
eliminar_movimiento = getattr(_db_manager, "eliminar_movimiento", None)

if actualizar_movimiento is None:
    def actualizar_movimiento(id_mov, data):
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, """UPDATE movimientos SET
                fecha_parte=?, modulo=?, accion=?, orden=?, nombre=?, aula=?, detalle=?
                WHERE id=?""",
                (
                    data.get('fecha_parte'), data.get('modulo'), data.get('accion'),
                    data.get('orden'), data.get('nombre'), data.get('aula'),
                    data.get('detalle', ''), id_mov
                ))
            conn.commit()

if eliminar_movimiento is None:
    def eliminar_movimiento(id_mov):
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, "DELETE FROM movimientos WHERE id=?", (id_mov,))
            conn.commit()


if not hasattr(_db_manager, "contar_usuarios"):
    import hashlib as _hashlib
    import hmac as _hmac
    import os as _os
    from datetime import datetime as _dt_auth

    def _auth_hash_password(password, salt=None):
        if salt is None:
            salt = _os.urandom(16).hex()
        digest = _hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), bytes.fromhex(salt), 120000)
        return salt, digest.hex()

    def _auth_ensure_table():
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, """CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                rol TEXT NOT NULL DEFAULT 'usuario',
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL
            )""")
            conn.commit()

    def _contar_usuarios():
        _auth_ensure_table()
        with _db_manager.get_db() as conn:
            cur = _db_manager.run(conn, "SELECT COUNT(*) AS total FROM usuarios")
            row = cur.fetchone()
            return int(row["total"] if isinstance(row, dict) else row[0])

    def _crear_usuario(usuario, password, rol="usuario", activo=True):
        _auth_ensure_table()
        usuario = str(usuario).strip().lower()
        salt, password_hash = _auth_hash_password(password)
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, """INSERT INTO usuarios (usuario, password_hash, salt, rol, activo, creado_en)
                VALUES (?, ?, ?, ?, ?, ?)""",
                (usuario, password_hash, salt, rol, 1 if activo else 0, _dt_auth.now().isoformat(timespec="seconds")))
            conn.commit()

    def _obtener_usuario(usuario):
        _auth_ensure_table()
        with _db_manager.get_db() as conn:
            cur = _db_manager.run(conn, "SELECT * FROM usuarios WHERE usuario=?", (str(usuario).strip().lower(),))
            row = cur.fetchone()
            return dict(row) if row else None

    def _autenticar_usuario(usuario, password):
        user = _obtener_usuario(usuario)
        if not user or int(user.get("activo", 0)) != 1:
            return None
        _, password_hash = _auth_hash_password(password, user["salt"])
        if not _hmac.compare_digest(password_hash, user["password_hash"]):
            return None
        return {"id": user["id"], "usuario": user["usuario"], "rol": user.get("rol", "usuario")}

    def _listar_usuarios():
        _auth_ensure_table()
        with _db_manager.get_db() as conn:
            cur = _db_manager.run(conn, "SELECT id, usuario, rol, activo, creado_en FROM usuarios ORDER BY usuario")
            return _db_manager.fetch_all(cur)

    def _actualizar_password_usuario(id_usuario, password):
        salt, password_hash = _auth_hash_password(password)
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, "UPDATE usuarios SET password_hash=?, salt=? WHERE id=?", (password_hash, salt, id_usuario))
            conn.commit()

    def _actualizar_estado_usuario(id_usuario, activo):
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, "UPDATE usuarios SET activo=? WHERE id=?", (1 if activo else 0, id_usuario))
            conn.commit()

    def _actualizar_rol_usuario(id_usuario, rol):
        rol = "admin" if str(rol).strip().lower() == "admin" else "usuario"
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, "UPDATE usuarios SET rol=? WHERE id=?", (rol, id_usuario))
            conn.commit()

    def _crear_o_actualizar_usuario_admin(usuario, password, solo_admin=False):
        usuario = str(usuario).strip().lower()
        user = _obtener_usuario(usuario)
        if user:
            _actualizar_password_usuario(user["id"], password)
            with _db_manager.get_db() as conn:
                _db_manager.run(conn, "UPDATE usuarios SET rol='admin', activo=1 WHERE id=?", (user["id"],))
                if solo_admin:
                    _db_manager.run(conn, "UPDATE usuarios SET rol='usuario' WHERE usuario<>?", (usuario,))
                conn.commit()
            return _obtener_usuario(usuario)
        _crear_usuario(usuario, password, rol="admin", activo=True)
        if solo_admin:
            with _db_manager.get_db() as conn:
                _db_manager.run(conn, "UPDATE usuarios SET rol='usuario' WHERE usuario<>?", (usuario,))
                conn.commit()
        return _obtener_usuario(usuario)

    def _eliminar_usuario(id_usuario):
        with _db_manager.get_db() as conn:
            _db_manager.run(conn, "DELETE FROM usuarios WHERE id=?", (id_usuario,))
            conn.commit()

    _db_manager.contar_usuarios = _contar_usuarios
    _db_manager.crear_usuario = _crear_usuario
    _db_manager.obtener_usuario = _obtener_usuario
    _db_manager.autenticar_usuario = _autenticar_usuario
    _db_manager.listar_usuarios = _listar_usuarios
    _db_manager.actualizar_password_usuario = _actualizar_password_usuario
    _db_manager.actualizar_estado_usuario = _actualizar_estado_usuario
    _db_manager.actualizar_rol_usuario = _actualizar_rol_usuario
    _db_manager.crear_o_actualizar_usuario_admin = _crear_o_actualizar_usuario_admin
    _db_manager.eliminar_usuario = _eliminar_usuario



def usuario_actual():
    return st.session_state.get("usuario_actual")


def es_admin():
    user = usuario_actual() or {}
    return user.get("rol") == "admin"


def cerrar_sesion():
    for key in ("usuario_actual", "autenticado"):
        if key in st.session_state:
            del st.session_state[key]


def obtener_secret_o_env(clave, default=None):
    valor = os.getenv(clave)
    if valor not in (None, ""):
        return valor
    try:
        return st.secrets.get(clave, default)
    except Exception:
        return default


def valor_verdadero(valor):
    return str(valor or "").strip().lower() in {"1", "true", "si", "sí", "yes", "y"}


def aplicar_recuperacion_admin():
    """Recupera/actualiza el administrador propietario usando Secrets o variables de entorno.

    Secrets admitidos:
    OWNER_ADMIN_ENABLED=true
    OWNER_ADMIN_USER="tu_usuario"
    OWNER_ADMIN_PASSWORD="tu_contraseña"
    OWNER_ADMIN_SOLO=true  # opcional: deja a los demás administradores como usuario
    """
    if not valor_verdadero(obtener_secret_o_env("OWNER_ADMIN_ENABLED", "false")):
        return

    usuario = str(obtener_secret_o_env("OWNER_ADMIN_USER", "")).strip()
    password = str(obtener_secret_o_env("OWNER_ADMIN_PASSWORD", "")).strip()
    solo_admin = valor_verdadero(obtener_secret_o_env("OWNER_ADMIN_SOLO", "false"))

    if not usuario or not password:
        st.warning("Recuperación admin activada, pero faltan OWNER_ADMIN_USER y/o OWNER_ADMIN_PASSWORD en Secrets.")
        return
    if len(password) < 6:
        st.warning("OWNER_ADMIN_PASSWORD debe tener al menos 6 caracteres.")
        return

    import hashlib as _hashlib
    firma_db = getattr(_db_manager, "firma_base_datos", lambda: "default")()
    firma_recuperacion = _hashlib.sha256(f"{firma_db}|{usuario}|{password}|{solo_admin}".encode("utf-8")).hexdigest()
    if st.session_state.get("admin_recovery_firma") == firma_recuperacion:
        return

    if hasattr(_db_manager, "crear_o_actualizar_usuario_admin"):
        _db_manager.crear_o_actualizar_usuario_admin(usuario, password, solo_admin=solo_admin)
        st.session_state.admin_recovery_applied = True
        st.session_state.admin_recovery_firma = firma_recuperacion


def requerir_login():
    total_usuarios = _db_manager.contar_usuarios()

    if total_usuarios == 0:
        col_l, col_c, col_r = st.columns([1, 1.25, 1])
        with col_c:
            logo_login_uri = asset_data_uri(APP_LOGO_FILE)
            if logo_login_uri:
                st.markdown(f'<div class="login-logo-wrap"><img class="login-logo" src="{logo_login_uri}" alt="Escuadron H"></div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="login-hero">
                <span class="login-badge">Acceso restringido</span>
                <h1 class="login-title">Escuadron H</h1>
                <p class="login-subtitle">No hay un administrador propietario configurado para esta base.</p>
            </div>
            <div class="login-card-note">Por seguridad, ya no se permite crear administradores desde el enlace publico. Configura OWNER_ADMIN_ENABLED, OWNER_ADMIN_USER y OWNER_ADMIN_PASSWORD en Secrets para recuperar el acceso.</div>
            """, unsafe_allow_html=True)
            st.error("Acceso bloqueado: falta configurar el administrador propietario.")
        st.stop()

    if not st.session_state.get("autenticado"):
        col_l, col_c, col_r = st.columns([1, 1.15, 1])
        with col_c:
            logo_login_uri = asset_data_uri(APP_LOGO_FILE)
            if logo_login_uri:
                st.markdown(f'<div class="login-logo-wrap"><img class="login-logo" src="{logo_login_uri}" alt="Escuadr?n H"></div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="login-hero">
                <span class="login-badge">Acceso restringido</span>
                <h1 class="login-title">Escuadrón H</h1>
                <p class="login-subtitle">Sistema de control de parte diario, novedades y reportes.</p>
            </div>
            """, unsafe_allow_html=True)
            with st.form("login_form"):
                usuario = st.text_input("Usuario")
                password = st.text_input("Contraseña", type="password")
                entrar = st.form_submit_button("Ingresar", type="primary", use_container_width=True)
        if entrar:
            user = _db_manager.autenticar_usuario(usuario, password)
            if user:
                st.session_state.autenticado = True
                st.session_state.usuario_actual = user
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")
        st.stop()

    with st.sidebar:
        if os.path.exists(APP_LOGO_FILE):
            st.image(APP_LOGO_FILE, use_container_width=True)
        user = usuario_actual() or {}
        st.markdown(f"""
        <div class="sidebar-card">
            <span>Sesión activa</span>
            <strong>{escape(str(user.get('usuario', '-')))}</strong>
            <small>Rol: {escape(str(user.get('rol', '-')))}</small>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Cerrar sesión", use_container_width=True):
            cerrar_sesion()
            st.rerun()




def obtener_query_param(nombre, default=""):
    try:
        valor = st.query_params.get(nombre, default)
        if isinstance(valor, list):
            return valor[0] if valor else default
        return valor if valor is not None else default
    except Exception:
        try:
            params = st.experimental_get_query_params()
            valor = params.get(nombre, [default])
            return valor[0] if isinstance(valor, list) else valor
        except Exception:
            return default


def formulario_plan_publico_solicitado():
    return str(obtener_query_param("form", "")).strip().lower() in {"plan", "plan_llamada", "llamada"}


def formulario_plan_token_valido():
    import hmac as _hmac
    if not valor_verdadero(obtener_secret_o_env("PLAN_FORM_ENABLED", "false")):
        return False, "El formulario público no está habilitado."
    token_cfg = str(obtener_secret_o_env("PLAN_FORM_TOKEN", "")).strip()
    token_url = str(obtener_query_param("token", "")).strip()
    if not token_cfg:
        return False, "Falta configurar PLAN_FORM_TOKEN en Secrets."
    if not token_url:
        return False, "Falta el token del formulario en el enlace."
    if not _hmac.compare_digest(token_cfg, token_url):
        return False, "Token inválido. Verifique el enlace recibido."
    return True, ""


def mostrar_formulario_plan_publico(df_personal):
    st.markdown("""
    <div class="login-hero">
        <span class="login-badge">Formulario de actualización</span>
        <h1 class="login-title">Plan de llamada y licencia</h1>
        <p class="login-subtitle">Completá solo los campos solicitados para licencia, traslado y contacto de emergencia.</p>
    </div>
    """, unsafe_allow_html=True)

    ok_token, msg_token = formulario_plan_token_valido()
    if not ok_token:
        st.error(msg_token)
        st.stop()

    st.info("Ingresá tu DNI o CE para validar tus datos ya cargados. El formulario no modifica la base de personal: solo agrega o actualiza los campos de licencia.")
    identificador = st.text_input("DNI o CE", placeholder="Escribí tu DNI o CE", key="plan_publico_identificador")

    if not identificador.strip():
        st.stop()

    ident = identificador.strip().upper().replace(" ", "")
    df_busqueda = df_personal.copy()
    df_busqueda["_DNI"] = df_busqueda["DNI"].astype(str).str.upper().str.replace(" ", "", regex=False)
    df_busqueda["_CE"] = df_busqueda["CE"].astype(str).str.upper().str.replace(" ", "", regex=False)
    encontrados = df_busqueda[(df_busqueda["_DNI"] == ident) | (df_busqueda["_CE"] == ident)]

    if encontrados.empty:
        st.warning("No se encontró personal con ese DNI/CE. Verificá el número o comunicate con administración.")
        st.stop()

    if len(encontrados) > 1:
        opciones = [f"{int(r['ORDEN_LIMP'])} - {r['NOMBRE_COMPLETO']} - {r['AULA']}" for _, r in encontrados.iterrows()]
        seleccionado = st.selectbox("Se encontró más de un registro. Seleccioná el correcto:", opciones)
        orden_sel = int(seleccionado.split(" - ", 1)[0])
        row = encontrados[encontrados["ORDEN_LIMP"].astype(int) == orden_sel].iloc[0]
    else:
        row = encontrados.iloc[0]

    orden = int(row["ORDEN_LIMP"])
    contacto = obtener_contacto(orden) or {}

    def limpio_base(valor):
        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass
        txt = str(valor or "").strip()
        return "" if txt.lower() == "nan" else txt

    nombre_base = limpio_base(row["NOMBRE_COMPLETO"]).upper()
    dni_base = limpio_base(row["DNI"])
    ce_base = limpio_base(row["CE"])
    aula_base = limpio_base(row["AULA"])

    st.success(f"Datos validados: {nombre_base} | Aula {aula_base} | Orden {orden}")
    st.caption("Los datos validados de apellido/nombre, DNI, CE y aula se toman de la base ya cargada y quedan bloqueados. Solo se completan los campos nuevos.")

    def idx_si_no(valor, defecto="NO"):
        valor = str(valor or defecto).strip().upper()
        return 0 if valor == "SI" else 1

    ctrans, cveh = st.columns(2)
    with ctrans:
        viaja_transporte = st.radio(
            "¿Viaja en transporte público?",
            ["SI", "NO"],
            index=idx_si_no(contacto.get("viaja_transporte_publico")),
            horizontal=True,
            key=f"pub_transporte_{orden}",
        )
    with cveh:
        viaja_vehiculo = st.radio(
            "¿Viaja en vehículo particular?",
            ["SI", "NO"],
            index=idx_si_no(contacto.get("viaja_vehiculo_particular")),
            horizontal=True,
            key=f"pub_vehiculo_{orden}",
        )

    with st.form("formulario_publico_plan_llamada"):
        st.markdown("### Datos validados de la base")
        st.text_input("Apellido y nombres", value=nombre_base, disabled=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.text_input("DNI", value=dni_base, disabled=True)
        with c2:
            st.text_input("CE", value=ce_base, disabled=True)
        with c3:
            st.text_input("Aula", value=aula_base, disabled=True)

        st.markdown("### Campos para completar")
        lugar_licencia = st.text_input("Lugar de licencia", value=contacto.get("lugar_licencia", ""), placeholder="Ciudad/localidad/provincia")
        direccion = st.text_area("Dirección", value=contacto.get("direccion", ""), placeholder="Dirección completa del lugar de licencia")
        c3, c4, c5 = st.columns(3)
        with c3:
            barrio = st.text_input("Barrio", value=contacto.get("barrio", ""), placeholder="Barrio")
        with c4:
            calle = st.text_input("Calle", value=contacto.get("calle", ""), placeholder="Calle")
        with c5:
            numero = st.text_input("Número", value=contacto.get("numero", ""), placeholder="Número")
        unidad_proxima_gn = st.text_input("Unidad más próxima de GN", value=contacto.get("unidad_proxima_gn", ""), placeholder="Unidad/Subunidad más cercana")

        if viaja_vehiculo == "SI":
            st.markdown("### Vehículo particular")
            cv1, cv2 = st.columns(2)
            with cv1:
                vehiculo_marca = st.text_input("Marca", value=contacto.get("vehiculo_marca", ""), placeholder="Ej: Toyota")
                vehiculo_dominio = st.text_input("Dominio", value=contacto.get("vehiculo_dominio", ""), placeholder="Patente/dominio")
            with cv2:
                vehiculo_modelo = st.text_input("Modelo", value=contacto.get("vehiculo_modelo", ""), placeholder="Ej: Hilux")
                vehiculo_titular = st.text_input("Titular del vehículo", value=contacto.get("vehiculo_titular", ""), placeholder="Nombre del titular")
        else:
            vehiculo_marca = ""
            vehiculo_modelo = ""
            vehiculo_dominio = ""
            vehiculo_titular = ""

        st.markdown("### Teléfonos")
        ct1, ct2 = st.columns(2)
        with ct1:
            telefono_particular = st.text_input("Teléfono particular", value=contacto.get("telefono_particular", ""), placeholder="Teléfono propio")
        with ct2:
            telefono_emergencia_licencia = st.text_input("Teléfono de emergencia para licencia", value=contacto.get("telefono_emergencia_licencia", ""), placeholder="Teléfono de emergencia")
        observaciones = st.text_area("Observaciones", value=contacto.get("observaciones", ""), placeholder="Observaciones opcionales")
        confirmar = st.checkbox("Confirmo que los datos cargados son correctos y están actualizados")
        enviar = st.form_submit_button("Guardar mis datos", type="primary", use_container_width=True)

    if enviar:
        faltantes = []
        if not lugar_licencia.strip(): faltantes.append("lugar de licencia")
        if not direccion.strip(): faltantes.append("dirección")
        if not barrio.strip(): faltantes.append("barrio")
        if not calle.strip(): faltantes.append("calle")
        if not numero.strip(): faltantes.append("número")
        if not unidad_proxima_gn.strip(): faltantes.append("unidad más próxima de GN")
        if not telefono_particular.strip(): faltantes.append("teléfono particular")
        if not telefono_emergencia_licencia.strip(): faltantes.append("teléfono de emergencia para licencia")
        if viaja_vehiculo == "SI":
            if not vehiculo_marca.strip(): faltantes.append("marca del vehículo")
            if not vehiculo_modelo.strip(): faltantes.append("modelo del vehículo")
            if not vehiculo_dominio.strip(): faltantes.append("dominio del vehículo")
            if not vehiculo_titular.strip(): faltantes.append("titular del vehículo")
        if not confirmar: faltantes.append("confirmación")

        if faltantes:
            st.error("Falta completar: " + ", ".join(faltantes) + ".")
        else:
            guardar_contacto({
                "orden": orden,
                "apellido_nombres": nombre_base,
                "dni": dni_base,
                "ce": ce_base,
                "viaja_transporte_publico": viaja_transporte,
                "viaja_vehiculo_particular": viaja_vehiculo,
                "vehiculo_marca": vehiculo_marca.strip().upper(),
                "vehiculo_modelo": vehiculo_modelo.strip().upper(),
                "vehiculo_dominio": vehiculo_dominio.strip().upper(),
                "vehiculo_titular": vehiculo_titular.strip().upper(),
                "lugar_licencia": lugar_licencia.strip().upper(),
                "direccion": direccion.strip().upper(),
                "barrio": barrio.strip().upper(),
                "calle": calle.strip().upper(),
                "numero": numero.strip().upper(),
                "unidad_proxima_gn": unidad_proxima_gn.strip().upper(),
                "telefono_particular": telefono_particular.strip(),
                "telefono_emergencia_licencia": telefono_emergencia_licencia.strip(),
                "observaciones": observaciones.strip().upper(),
                "actualizado_en": ahora_local().isoformat(timespec="seconds"),
            })
            try:
                registrar_movimiento(
                    ahora_local().date().isoformat(),
                    "Plan de llamada",
                    "FORMULARIO LICENCIA",
                    orden,
                    nombre_base,
                    aula_base,
                    "Datos de licencia actualizados por formulario público",
                )
            except Exception:
                pass
            st.success("✅ Tus datos fueron guardados correctamente. Muchas gracias.")
            st.info("Ya podés cerrar esta pantalla.")
            st.stop()


def formulario_dinamico_publico_solicitado():
    return bool(str(obtener_query_param("formulario", "")).strip())


def _texto_limpio(valor):
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass
    txt = str(valor or "").strip()
    return "" if txt.lower() == "nan" else txt


def _solo_digitos(valor):
    return re.sub(r"\D+", "", str(valor or ""))


def _mayus(valor):
    return str(valor or "").strip().upper()


def _dominio(valor):
    return re.sub(r"\s+", "", _mayus(valor))


def _validar_numero(nombre, valor, obligatorio=True):
    texto = str(valor or "").strip()
    if not texto and not obligatorio:
        return ""
    if not texto:
        return f"Falta completar {nombre}."
    if not texto.isdigit():
        return f"{nombre} debe tener solo numeros, sin puntos, letras ni espacios."
    return ""


TIPOS_CAMPOS_FORMULARIO = {
    "Texto corto": "texto",
    "Texto largo": "texto_largo",
    "Numero": "numero",
    "Telefono": "telefono",
    "DNI": "dni",
    "CE": "ce",
    "SI/NO": "si_no",
    "Lista desplegable": "lista",
    "Fecha": "fecha",
    "Hora": "hora",
}


def clave_campo_desde_etiqueta(etiqueta):
    base = unicodedata.normalize("NFKD", str(etiqueta or "")).encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^a-zA-Z0-9]+", "_", base).strip("_").lower()
    return base or "campo"


def normalizar_campo_formulario(campo):
    campo = dict(campo or {})
    etiqueta = str(campo.get("etiqueta") or campo.get("label") or "").strip()
    clave = str(campo.get("clave") or campo.get("name") or clave_campo_desde_etiqueta(etiqueta)).strip().lower()
    tipo = str(campo.get("tipo") or "texto").strip().lower()
    opciones = campo.get("opciones") or []
    if isinstance(opciones, str):
        opciones = [o.strip().upper() for o in opciones.split(",") if o.strip()]
    if tipo == "si_no":
        opciones = ["SI", "NO"]
    return {
        "clave": clave_campo_desde_etiqueta(clave),
        "etiqueta": etiqueta or clave.replace("_", " ").title(),
        "tipo": tipo if tipo in TIPOS_CAMPOS_FORMULARIO.values() else "texto",
        "obligatorio": bool(campo.get("obligatorio", False)),
        "mayusculas": bool(campo.get("mayusculas", tipo in {"texto", "texto_largo", "lista", "si_no"})),
        "sin_puntos": bool(campo.get("sin_puntos", tipo in {"dni", "telefono"})),
        "solo_numeros": bool(campo.get("solo_numeros", tipo in {"numero", "dni", "telefono", "ce"})),
        "opciones": opciones,
    }


def campos_formulario_licencia_base():
    return [
        {"clave": "viaja_transporte_publico", "etiqueta": "Viaja en transporte publico", "tipo": "si_no", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": ["SI", "NO"]},
        {"clave": "viaja_vehiculo_particular", "etiqueta": "Viaja en vehiculo particular", "tipo": "si_no", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": ["SI", "NO"]},
        {"clave": "lugar_licencia", "etiqueta": "Lugar de licencia", "tipo": "texto", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "direccion", "etiqueta": "Direccion", "tipo": "texto", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "barrio", "etiqueta": "Barrio", "tipo": "texto", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "calle", "etiqueta": "Calle", "tipo": "texto", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "numero", "etiqueta": "Numero", "tipo": "numero", "obligatorio": True, "mayusculas": False, "sin_puntos": True, "solo_numeros": True, "opciones": []},
        {"clave": "unidad_proxima_gn", "etiqueta": "Unidad GN proxima", "tipo": "texto", "obligatorio": True, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "telefono_particular", "etiqueta": "Telefono particular", "tipo": "telefono", "obligatorio": True, "mayusculas": False, "sin_puntos": True, "solo_numeros": True, "opciones": []},
        {"clave": "telefono_emergencia", "etiqueta": "Telefono emergencia", "tipo": "telefono", "obligatorio": True, "mayusculas": False, "sin_puntos": True, "solo_numeros": True, "opciones": []},
        {"clave": "vehiculo_marca", "etiqueta": "Marca del vehiculo", "tipo": "texto", "obligatorio": False, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "vehiculo_modelo", "etiqueta": "Modelo del vehiculo", "tipo": "texto", "obligatorio": False, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "vehiculo_dominio", "etiqueta": "Dominio del vehiculo", "tipo": "texto", "obligatorio": False, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "vehiculo_titular", "etiqueta": "Titular del vehiculo", "tipo": "texto", "obligatorio": False, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
        {"clave": "observaciones", "etiqueta": "Observaciones", "tipo": "texto_largo", "obligatorio": False, "mayusculas": True, "sin_puntos": False, "solo_numeros": False, "opciones": []},
    ]


def obtener_campos_formulario(formulario):
    reglas = formulario.get("reglas") or {}
    if isinstance(reglas, list):
        campos = reglas
    elif isinstance(reglas, dict) and isinstance(reglas.get("campos"), list):
        campos = reglas.get("campos", [])
    elif isinstance(reglas, dict) and reglas.get("tipo") == "licencia_contacto":
        campos = campos_formulario_licencia_base()
    else:
        campos = []
    return [normalizar_campo_formulario(c) for c in campos]


def reglas_desde_campos(campos):
    return {"version": 2, "campos": [normalizar_campo_formulario(c) for c in campos]}


def validar_y_normalizar_valor_campo(campo, valor):
    campo = normalizar_campo_formulario(campo)
    tipo = campo["tipo"]
    if valor is None:
        texto = ""
    elif tipo == "fecha":
        texto = valor.isoformat() if hasattr(valor, "isoformat") else str(valor)
    elif tipo == "hora":
        texto = valor.strftime("%H:%M") if hasattr(valor, "strftime") else str(valor)
    else:
        texto = str(valor).strip()

    if campo["obligatorio"] and not texto:
        return "", f"Falta completar {campo['etiqueta']}."
    if not texto:
        return "", ""
    if campo["sin_puntos"] and "." in texto:
        return "", f"{campo['etiqueta']} no permite puntos."
    if tipo in {"dni", "telefono"}:
        campo["solo_numeros"] = True
        campo["sin_puntos"] = True
    if tipo == "si_no":
        texto = texto.upper()
        if texto not in {"SI", "NO"}:
            return "", f"{campo['etiqueta']} solo permite SI o NO."
    if campo["solo_numeros"] and not texto.isdigit():
        return "", f"{campo['etiqueta']} debe tener solo numeros, sin puntos, comas, espacios ni guiones."
    if tipo == "lista":
        opciones = [str(o).strip().upper() if campo["mayusculas"] else str(o).strip() for o in campo.get("opciones", [])]
        comparado = texto.upper() if campo["mayusculas"] else texto
        if opciones and comparado not in opciones:
            return "", f"{campo['etiqueta']} debe ser una opcion valida."
    if campo["mayusculas"]:
        texto = texto.upper()
    return texto, ""


def renderizar_campo_formulario(campo, valor_actual, key_prefix):
    campo = normalizar_campo_formulario(campo)
    clave = campo["clave"]
    etiqueta = campo["etiqueta"] + (" *" if campo["obligatorio"] else "")
    tipo = campo["tipo"]
    key = f"{key_prefix}_{clave}"
    valor_actual = "" if valor_actual is None else valor_actual
    if tipo == "texto_largo":
        return st.text_area(etiqueta, value=str(valor_actual), key=key)
    if tipo in {"numero", "telefono", "dni", "ce"}:
        return st.text_input(etiqueta, value=str(valor_actual), key=key)
    if tipo == "si_no":
        opciones = ["SI", "NO"]
        idx = 0 if str(valor_actual).upper() == "SI" else 1
        return st.radio(etiqueta, opciones, index=idx, horizontal=True, key=key)
    if tipo == "lista":
        opciones = [str(o).strip().upper() if campo["mayusculas"] else str(o).strip() for o in campo.get("opciones", []) if str(o).strip()]
        if not opciones:
            opciones = [""]
        idx = opciones.index(valor_actual) if valor_actual in opciones else 0
        return st.selectbox(etiqueta, opciones, index=idx, key=key)
    if tipo == "fecha":
        fecha_default = None
        try:
            fecha_default = datetime.fromisoformat(str(valor_actual)).date() if valor_actual else None
        except Exception:
            fecha_default = None
        return st.date_input(etiqueta, value=fecha_default, key=key)
    if tipo == "hora":
        hora_default = None
        try:
            hora_default = datetime.strptime(str(valor_actual), "%H:%M").time() if valor_actual else None
        except Exception:
            hora_default = None
        return st.time_input(etiqueta, value=hora_default, key=key)
    return st.text_input(etiqueta, value=str(valor_actual), key=key)


def link_publico_formulario(formulario):
    slug = str(formulario.get("slug", "")).strip()
    token = str(formulario.get("token", "")).strip()
    base = str(obtener_secret_o_env("APP_PUBLIC_URL", "") or "").strip().rstrip("/")
    final = f"?formulario={slug}&token={token}"
    return f"{base}/{final}" if base else final


def buscar_personal_por_identificador(df_personal, identificador):
    ident = str(identificador or "").strip().upper().replace(" ", "")
    if not ident:
        return df_personal.iloc[0:0]
    df_busqueda = df_personal.copy()
    df_busqueda["_DNI"] = df_busqueda["DNI"].astype(str).str.upper().str.replace(" ", "", regex=False).str.replace(".", "", regex=False).str.replace("-", "", regex=False)
    df_busqueda["_CE"] = df_busqueda["CE"].astype(str).str.upper().str.replace(" ", "", regex=False).str.replace(".", "", regex=False).str.replace("-", "", regex=False)
    return df_busqueda[(df_busqueda["_DNI"] == ident) | (df_busqueda["_CE"] == ident)]


def mostrar_formulario_dinamico_publico(df_personal):
    slug = str(obtener_query_param("formulario", "")).strip().lower()
    token = str(obtener_query_param("token", "")).strip()
    formulario = _db_manager.obtener_formulario_por_slug_y_token(slug, token)

    st.markdown("""
    <div class="login-hero">
        <span class="login-badge">Formulario seguro</span>
        <h1 class="login-title">Escuadrón H</h1>
        <p class="login-subtitle">Completá los datos solicitados. Se valida tu DNI o CE contra la base del escuadrón.</p>
    </div>
    """, unsafe_allow_html=True)

    if not formulario:
        st.error("El enlace del formulario no es valido o el token no corresponde.")
        st.stop()
    if int(formulario.get("activo", 0)) != 1:
        st.warning("Este formulario se encuentra desactivado.")
        st.stop()

    st.subheader(str(formulario.get("nombre", "Formulario")).strip())
    if formulario.get("descripcion"):
        st.caption(str(formulario.get("descripcion")))

    campos = obtener_campos_formulario(formulario)
    if not campos:
        st.warning("Este formulario todavía no tiene campos configurados.\nAgregue campos desde el panel de administración.")
        st.stop()

    identificador = st.text_input("DNI o CE", placeholder="Escribi tu DNI o CE sin puntos", key=f"formpub_ident_{slug}")
    if not identificador.strip():
        st.stop()

    encontrados = buscar_personal_por_identificador(df_personal, identificador)
    if encontrados.empty:
        st.warning("No se encontró personal con ese DNI/CE. Verificá el número o comunicate con administración.")
        st.stop()

    if len(encontrados) > 1:
        opciones = [f"{int(r['ORDEN_LIMP'])} - {r['NOMBRE_COMPLETO']} - {r['AULA']}" for _, r in encontrados.iterrows()]
        seleccionado = st.selectbox("Se encontró más de un registro. Seleccioná el correcto:", opciones)
        orden_sel = int(seleccionado.split(" - ", 1)[0])
        row = encontrados[encontrados["ORDEN_LIMP"].astype(int) == orden_sel].iloc[0]
    else:
        row = encontrados.iloc[0]

    orden = int(row["ORDEN_LIMP"])
    respuestas_previas = _db_manager.obtener_respuestas_formulario(formulario["id"])
    previa = next((r for r in respuestas_previas if int(r.get("orden", 0)) == orden), None)
    datos_previos = previa.get("datos", {}) if previa else {}

    nombre_base = _texto_limpio(row["NOMBRE_COMPLETO"]).upper()
    dni_base = _solo_digitos(row["DNI"])
    ce_base = _solo_digitos(row["CE"])
    aula_base = _texto_limpio(row["AULA"]).upper()

    st.success(f"Datos validados: {nombre_base} | Aula {aula_base} | Orden {orden}")
    st.caption("Apellido/nombres, DNI, CE, Aula y Orden se toman de alumnos.csv y quedan bloqueados.")

    with st.form(f"formulario_dinamico_{formulario['id']}_{orden}"):
        st.markdown("### Datos de base")
        st.text_input("Apellido y nombres", value=nombre_base, disabled=True)
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.text_input("Orden", value=str(orden), disabled=True)
        with c2:
            st.text_input("DNI", value=dni_base, disabled=True)
        with c3:
            st.text_input("CE", value=ce_base, disabled=True)
        with c4:
            st.text_input("Aula", value=aula_base, disabled=True)

        st.markdown("### Datos para completar")
        valores_ingresados = {}
        for campo in campos:
            valores_ingresados[campo["clave"]] = renderizar_campo_formulario(campo, datos_previos.get(campo["clave"], ""), f"formpub_{formulario['id']}_{orden}")
        confirmar = st.checkbox("Confirmo que los datos cargados son correctos")
        enviar = st.form_submit_button("Guardar respuesta", type="primary", use_container_width=True)

    if enviar:
        errores = []
        datos = {}
        for campo in campos:
            valor, error = validar_y_normalizar_valor_campo(campo, valores_ingresados.get(campo["clave"]))
            if error:
                errores.append(error)
            datos[campo["clave"]] = valor
        if not confirmar:
            errores.append("Falta confirmar que los datos son correctos.")

        if errores:
            st.error(" ".join(errores))
        else:
            _db_manager.guardar_respuesta_formulario(formulario["id"], orden, nombre_base, dni_base, ce_base, aula_base, datos)
            st.success("Tus datos fueron guardados correctamente. Muchas gracias.")
            st.info("Ya podés cerrar esta pantalla.")
            st.stop()


def preparar_control_formulario(df_personal, respuestas, campos):
    respuestas_por_orden = {int(r.get("orden", 0)): r for r in respuestas if str(r.get("orden", "")).strip()}
    completaron = []
    no_completaron = []
    for _, row in df_personal.iterrows():
        orden = int(row["ORDEN_LIMP"])
        base = {
            "Orden": orden,
            "Apellido y nombres": _texto_limpio(row["NOMBRE_COMPLETO"]),
            "DNI": _solo_digitos(row["DNI"]),
            "CE": _solo_digitos(row["CE"]),
            "Aula": _texto_limpio(row["AULA"]),
            "Curso": _texto_limpio(row["GRADO"]),
        }
        resp = respuestas_por_orden.get(orden)
        if resp:
            datos = resp.get("datos", {}) or {}
            completo = dict(base)
            completo["Fecha de carga"] = resp.get("creado_en", "")
            completo["Fecha de actualización"] = resp.get("actualizado_en", "")
            for campo in campos:
                completo[campo["etiqueta"]] = datos.get(campo["clave"], "")
            completaron.append(completo)
        else:
            no_completaron.append(base)
    return pd.DataFrame(completaron), pd.DataFrame(no_completaron)


def aplicar_filtros_control_formulario(df_base, texto="", aula="Todas", curso="Todos"):
    if df_base.empty:
        return df_base
    filtrado = df_base.copy()
    if aula and aula != "Todas" and "Aula" in filtrado.columns:
        filtrado = filtrado[filtrado["Aula"].astype(str) == aula]
    if curso and curso != "Todos" and "Curso" in filtrado.columns:
        filtrado = filtrado[filtrado["Curso"].astype(str) == curso]
    q = str(texto or "").strip().upper()
    if q:
        cols = [c for c in ["Apellido y nombres", "DNI", "CE", "Orden"] if c in filtrado.columns]
        mask = pd.Series(False, index=filtrado.index)
        for col in cols:
            mask = mask | filtrado[col].astype(str).str.upper().str.contains(q, na=False)
        filtrado = filtrado[mask]
    return filtrado


def excel_formulario_control(formulario, df_completaron, df_no_completaron):
    from openpyxl.styles import Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws_res = wb.active
    ws_res.title = "RESUMEN"
    ws_comp = wb.create_sheet("COMPLETARON")
    ws_pend = wb.create_sheet("NO_COMPLETARON")

    total = len(df_completaron) + len(df_no_completaron)
    avance = round((len(df_completaron) / total) * 100, 2) if total else 0
    ultima = ""
    if not df_completaron.empty and "Fecha de actualización" in df_completaron.columns:
        ultima = str(df_completaron["Fecha de actualización"].max())

    resumen = [
        ("Formulario", formulario.get("nombre", "")),
        ("Slug", formulario.get("slug", "")),
        ("Total esperados", total),
        ("Completaron", len(df_completaron)),
        ("No completaron", len(df_no_completaron)),
        ("Avance", f"{avance}%"),
        ("Ultima carga", ultima),
    ]

    fill_titulo = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
    fill_header = PatternFill(start_color=EXCEL_OLIVE, end_color=EXCEL_OLIVE, fill_type="solid")
    fill_alt = PatternFill(start_color=EXCEL_OLIVE_ROW, end_color=EXCEL_OLIVE_ROW, fill_type="solid")
    border = Border(left=Side(style="thin", color=EXCEL_BORDER), right=Side(style="thin", color=EXCEL_BORDER), top=Side(style="thin", color=EXCEL_BORDER), bottom=Side(style="thin", color=EXCEL_BORDER))

    ws_res.merge_cells("A1:B1")
    ws_res["A1"] = 'ESCUADRON H "Cabo Marcelo Godoy"'
    ws_res["A1"].font = excel_font(bold=True, size=14, color=EXCEL_WHITE)
    ws_res["A1"].fill = fill_titulo
    ws_res["A1"].alignment = Alignment(horizontal="center")
    ws_res["A2"] = "Resumen de formulario"
    ws_res["A2"].font = excel_font(bold=True, color=EXCEL_TEXT_DARK)
    for idx, (clave, valor) in enumerate(resumen, 4):
        ws_res.cell(idx, 1, clave)
        ws_res.cell(idx, 2, valor)
        for col in (1, 2):
            ws_res.cell(idx, col).border = border
    ws_res.column_dimensions["A"].width = 24
    ws_res.column_dimensions["B"].width = 44

    def escribir_tabla(ws, df_tabla):
        if df_tabla.empty:
            ws["A1"] = "Sin registros"
            return
        for col_idx, col_name in enumerate(df_tabla.columns, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE)
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row_idx, (_, row) in enumerate(df_tabla.iterrows(), 2):
            for col_idx, col_name in enumerate(df_tabla.columns, 1):
                cell = ws.cell(row_idx, col_idx, row.get(col_name, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = fill_alt
        for col_idx, col_name in enumerate(df_tabla.columns, 1):
            width = min(max(len(str(col_name)) + 4, 14), 32)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    escribir_tabla(ws_comp, df_completaron)
    escribir_tabla(ws_pend, df_no_completaron)
    return excel_bytes(wb)


def panel_formularios_dinamicos(df_personal):
    st.markdown("### Formularios")
    st.caption("Crea enlaces publicos con token, controla respuestas y exporta el resumen.")

    pin_cfg = str(obtener_secret_o_env("FORM_ADMIN_PIN", "") or "").strip()
    if not pin_cfg:
        st.warning("Falta configurar FORM_ADMIN_PIN en Secrets.")
        return

    pin_ingresado = st.text_input("PIN de formularios", type="password", key="form_admin_pin")
    if not pin_ingresado:
        st.info("Ingresá el PIN para administrar formularios.")
        return
    if not hmac.compare_digest(str(pin_ingresado), pin_cfg):
        st.error("PIN incorrecto.")
        return

    info_bd_form = getattr(_db_manager, "info_base_datos", lambda: {"persistente": False})()
    if not info_bd_form.get("persistente"):
        st.warning("Base local SQLite: sirve para pruebas. Para uso real multiusuario conviene DATABASE_URL en Secrets.")

    with st.expander("Crear formulario", expanded=False):
        with st.form("crear_formulario_dinamico"):
            nombre = st.text_input("Nombre interno", placeholder="Ej: Licencia invierno 2026")
            slug = st.text_input("Slug / codigo", placeholder="licencia26").strip().lower()
            descripcion = st.text_area("Descripcion opcional", placeholder="Texto breve que vera el aspirante")
            activo = st.toggle("Activo", value=True)
            crear = st.form_submit_button("Crear formulario", type="primary", use_container_width=True)
        if crear:
            slug_ok = re.fullmatch(r"[a-z0-9_-]{3,40}", slug or "")
            if not nombre.strip() or not slug_ok:
                st.error("Completá nombre y un slug válido: 3 a 40 caracteres, letras, números, guion o guion bajo.")
            elif _db_manager.obtener_formulario_por_slug(slug):
                st.error("Ya existe un formulario con ese slug.")
            else:
                token = py_secrets.token_urlsafe(24)
                _db_manager.crear_formulario(slug, nombre, descripcion, token, activo, reglas_desde_campos([]))
                st.success("Formulario creado.")
                st.rerun()

    formularios = _db_manager.listar_formularios()
    if not formularios:
        st.info("Todavía no hay formularios creados.")
        return

    opciones = {f"{f['nombre']} ({f['slug']})": f for f in formularios}
    etiqueta = st.selectbox("Formulario", list(opciones.keys()), key="form_admin_selector")
    formulario = opciones[etiqueta]

    c_link, c_estado = st.columns([3, 1])
    with c_link:
        st.text_input("Link publico", value=link_publico_formulario(formulario), key=f"link_form_{formulario['id']}")
        st.caption("Si no configuraste APP_PUBLIC_URL, copiá este final y pegalo al final de la URL de Streamlit.")
    with c_estado:
        st.metric("Estado", "Activo" if int(formulario.get("activo", 0)) == 1 else "Inactivo")

    with st.expander("Editar formulario", expanded=False):
        with st.form(f"editar_formulario_{formulario['id']}"):
            nuevo_nombre = st.text_input("Nombre", value=formulario.get("nombre", ""))
            nuevo_slug = st.text_input("Slug", value=formulario.get("slug", ""))
            nueva_desc = st.text_area("Descripcion", value=formulario.get("descripcion", "") or "")
            nuevo_activo = st.toggle("Activo", value=int(formulario.get("activo", 0)) == 1)
            guardar = st.form_submit_button("Guardar cambios", use_container_width=True)
        if guardar:
            nuevo_slug_norm = str(nuevo_slug or "").strip().lower()
            slug_ok = re.fullmatch(r"[a-z0-9_-]{3,40}", nuevo_slug_norm or "")
            otro = _db_manager.obtener_formulario_por_slug(nuevo_slug_norm)
            if not nuevo_nombre.strip() or not slug_ok:
                st.error("Nombre y slug son obligatorios.")
            elif otro and int(otro["id"]) != int(formulario["id"]):
                st.error("Ese slug ya está usado por otro formulario.")
            else:
                _db_manager.actualizar_formulario(formulario["id"], nombre=nuevo_nombre, slug=nuevo_slug_norm, descripcion=nueva_desc, activo=nuevo_activo)
                st.success("Formulario actualizado.")
                st.rerun()

    campos = obtener_campos_formulario(formulario)
    with st.expander("Campos del formulario", expanded=True):
        if not campos:
            st.info("Este formulario todavía no tiene campos configurados.\nAgregue campos desde el panel de administración.")

        for idx, campo in enumerate(campos):
            titulo = f"{idx + 1}. {campo['etiqueta']} ({campo['tipo']})"
            with st.container(border=True):
                st.markdown(f"**{titulo}**")
                e1, e2, e3 = st.columns([2, 1.4, 1])
                with e1:
                    etiqueta_c = st.text_input("Etiqueta", value=campo["etiqueta"], key=f"campo_etq_{formulario['id']}_{idx}")
                    clave_c = st.text_input("Clave interna", value=campo["clave"], key=f"campo_clave_{formulario['id']}_{idx}")
                    opciones_c = st.text_input("Opciones (separadas por coma)", value=", ".join(campo.get("opciones", [])), key=f"campo_opc_{formulario['id']}_{idx}")
                with e2:
                    tipo_label = next((k for k, v in TIPOS_CAMPOS_FORMULARIO.items() if v == campo["tipo"]), "Texto corto")
                    tipo_c = st.selectbox("Tipo", list(TIPOS_CAMPOS_FORMULARIO.keys()), index=list(TIPOS_CAMPOS_FORMULARIO.keys()).index(tipo_label), key=f"campo_tipo_{formulario['id']}_{idx}")
                    obligatorio_c = st.toggle("Obligatorio", value=campo["obligatorio"], key=f"campo_obl_{formulario['id']}_{idx}")
                    mayus_c = st.toggle("Mayusculas", value=campo["mayusculas"], key=f"campo_may_{formulario['id']}_{idx}")
                with e3:
                    sin_puntos_c = st.toggle("Sin puntos", value=campo["sin_puntos"], key=f"campo_pun_{formulario['id']}_{idx}")
                    solo_numeros_c = st.toggle("Solo numeros", value=campo["solo_numeros"], key=f"campo_num_{formulario['id']}_{idx}")
                    col_a, col_b = st.columns(2)
                    with col_a:
                        if st.button("Subir", key=f"campo_up_{formulario['id']}_{idx}", disabled=idx == 0, use_container_width=True):
                            campos[idx - 1], campos[idx] = campos[idx], campos[idx - 1]
                            _db_manager.actualizar_formulario(formulario["id"], reglas_json=reglas_desde_campos(campos))
                            st.rerun()
                    with col_b:
                        if st.button("Bajar", key=f"campo_down_{formulario['id']}_{idx}", disabled=idx == len(campos) - 1, use_container_width=True):
                            campos[idx + 1], campos[idx] = campos[idx], campos[idx + 1]
                            _db_manager.actualizar_formulario(formulario["id"], reglas_json=reglas_desde_campos(campos))
                            st.rerun()
                    if st.button("Eliminar", key=f"campo_del_{formulario['id']}_{idx}", use_container_width=True):
                        nuevos = [c for i, c in enumerate(campos) if i != idx]
                        _db_manager.actualizar_formulario(formulario["id"], reglas_json=reglas_desde_campos(nuevos))
                        st.rerun()
                if st.button("Guardar campo", key=f"campo_save_{formulario['id']}_{idx}", use_container_width=True):
                    campos[idx] = normalizar_campo_formulario({
                        "etiqueta": etiqueta_c,
                        "clave": clave_c,
                        "tipo": TIPOS_CAMPOS_FORMULARIO[tipo_c],
                        "obligatorio": obligatorio_c,
                        "mayusculas": mayus_c,
                        "sin_puntos": sin_puntos_c,
                        "solo_numeros": solo_numeros_c,
                        "opciones": opciones_c,
                    })
                    _db_manager.actualizar_formulario(formulario["id"], reglas_json=reglas_desde_campos(campos))
                    st.success("Campo actualizado.")
                    st.rerun()

        st.markdown("#### Agregar campo")
        with st.form(f"agregar_campo_{formulario['id']}"):
            n1, n2 = st.columns(2)
            with n1:
                nueva_etiqueta = st.text_input("Etiqueta del campo", placeholder="Lugar de licencia")
                nueva_clave = st.text_input("Clave interna", placeholder="Se genera si queda vacio")
                nuevo_tipo = st.selectbox("Tipo", list(TIPOS_CAMPOS_FORMULARIO.keys()))
            with n2:
                nuevo_obl = st.toggle("Obligatorio", value=True)
                nuevo_may = st.toggle("Convertir a mayusculas", value=True)
                nuevo_sin_puntos = st.toggle("No permitir puntos", value=False)
                nuevo_solo_nums = st.toggle("Solo numeros", value=False)
                nuevas_opciones = st.text_input("Opciones si corresponde", placeholder="SI, NO, OTRA")
            agregar_campo = st.form_submit_button("Agregar campo", type="primary", use_container_width=True)
        if agregar_campo:
            if not nueva_etiqueta.strip():
                st.error("La etiqueta del campo es obligatoria.")
            else:
                clave_final = nueva_clave.strip() or clave_campo_desde_etiqueta(nueva_etiqueta)
                claves_existentes = {c["clave"] for c in campos}
                if clave_campo_desde_etiqueta(clave_final) in claves_existentes:
                    st.error("Ya existe un campo con esa clave interna.")
                else:
                    campos.append(normalizar_campo_formulario({
                        "etiqueta": nueva_etiqueta,
                        "clave": clave_final,
                        "tipo": TIPOS_CAMPOS_FORMULARIO[nuevo_tipo],
                        "obligatorio": nuevo_obl,
                        "mayusculas": nuevo_may,
                        "sin_puntos": nuevo_sin_puntos,
                        "solo_numeros": nuevo_solo_nums,
                        "opciones": nuevas_opciones,
                    }))
                    _db_manager.actualizar_formulario(formulario["id"], reglas_json=reglas_desde_campos(campos))
                    st.success("Campo agregado.")
                    st.rerun()

    respuestas = _db_manager.obtener_respuestas_formulario(formulario["id"])
    df_completaron, df_no_completaron = preparar_control_formulario(df_personal, respuestas, campos)
    total = len(df_personal)
    avance = round((len(df_completaron) / total) * 100, 1) if total else 0
    ultima = "-"
    if not df_completaron.empty and "Fecha de actualización" in df_completaron.columns:
        ultima = str(df_completaron["Fecha de actualización"].max())

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Esperados", total)
    m2.metric("Completaron", len(df_completaron))
    m3.metric("No completaron", len(df_no_completaron))
    m4.metric("Avance", f"{avance}%")
    st.caption(f"Ultima carga: {ultima}")

    f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
    with f1:
        buscar = st.text_input("Buscar apellido, DNI, CE u orden", key=f"form_buscar_{formulario['id']}")
    with f2:
        aulas = ["Todas"] + sorted(df_personal["AULA"].astype(str).unique().tolist())
        aula = st.selectbox("Aula", aulas, key=f"form_aula_{formulario['id']}")
    with f3:
        cursos = ["Todos"] + sorted(df_personal["GRADO"].astype(str).unique().tolist())
        curso = st.selectbox("Curso", cursos, key=f"form_curso_{formulario['id']}")
    with f4:
        vista = st.selectbox("Vista", ["Todos", "Completaron", "Pendientes"], key=f"form_vista_{formulario['id']}")

    comp_filtrado = aplicar_filtros_control_formulario(df_completaron, buscar, aula, curso)
    pend_filtrado = aplicar_filtros_control_formulario(df_no_completaron, buscar, aula, curso)

    if vista in {"Todos", "Completaron"}:
        st.markdown("#### Completaron")
        st.dataframe(comp_filtrado, use_container_width=True, hide_index=True)
    if vista in {"Todos", "Pendientes"}:
        st.markdown("#### No completaron")
        st.dataframe(pend_filtrado, use_container_width=True, hide_index=True)

    excel_data = excel_formulario_control(formulario, comp_filtrado, pend_filtrado)
    st.download_button(
        "Descargar Excel del formulario",
        data=excel_data,
        file_name=f"FORMULARIO_{formulario.get('slug', 'formulario').upper()}_{ahora_local().strftime('%d%m%Y_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def panel_admin_usuarios():
    st.subheader("👥 Administración de usuarios")
    st.caption("Crea y administra los usuarios autorizados para ingresar a la plataforma.")

    with st.form("crear_usuario_form"):
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            nuevo_usuario = st.text_input("Nuevo usuario")
        with c2:
            nueva_password = st.text_input("Contrasena inicial", type="password")
        with c3:
            nuevo_rol = st.selectbox("Rol", ["usuario", "admin"])
        crear_usuario_btn = st.form_submit_button("Crear usuario", type="primary", use_container_width=True)
    if crear_usuario_btn:
        if not nuevo_usuario.strip() or not nueva_password:
            st.error("Completa usuario y contrasena.")
        elif len(nueva_password) < 6:
            st.error("La contrasena debe tener al menos 6 caracteres.")
        else:
            try:
                _db_manager.crear_usuario(nuevo_usuario, nueva_password, rol=nuevo_rol, activo=True)
                st.success("Usuario creado.")
                st.rerun()
            except Exception as e:
                st.error(f"No se pudo crear el usuario: {e}")

    usuarios = _db_manager.listar_usuarios()
    if not usuarios:
        st.info("No hay usuarios cargados.")
        return

    st.divider()
    for user in usuarios:
        with st.container(border=True):
            c_info, c_rol, c_estado, c_pass, c_del = st.columns([2.4, 1.2, 1.2, 2, 1])
            with c_info:
                estado = "Activo" if int(user.get("activo", 0)) == 1 else "Inactivo"
                st.markdown(f"**{user['usuario']}**")
                st.caption(f"Rol actual: {user.get('rol', 'usuario')} | {estado} | Creado: {user.get('creado_en', '-')}")
            with c_rol:
                rol_actual = user.get("rol", "usuario")
                opciones_rol = ["usuario", "admin"]
                idx_rol = opciones_rol.index(rol_actual) if rol_actual in opciones_rol else 0
                es_mi_usuario = user.get("usuario") == (usuario_actual() or {}).get("usuario")
                nuevo_rol = st.selectbox("Rol", opciones_rol, index=idx_rol, key=f"usr_rol_{user['id']}", disabled=es_mi_usuario)
                if nuevo_rol != rol_actual and hasattr(_db_manager, "actualizar_rol_usuario"):
                    _db_manager.actualizar_rol_usuario(user["id"], nuevo_rol)
                    st.rerun()
            with c_estado:
                activo = int(user.get("activo", 0)) == 1
                nuevo_estado = st.toggle("Activo", value=activo, key=f"usr_activo_{user['id']}")
                if nuevo_estado != activo:
                    _db_manager.actualizar_estado_usuario(user["id"], nuevo_estado)
                    st.rerun()
            with c_pass:
                nueva = st.text_input("Nueva contrasena", type="password", key=f"usr_pass_{user['id']}")
                if st.button("Cambiar", key=f"usr_cambiar_{user['id']}", use_container_width=True):
                    if len(nueva) < 6:
                        st.warning("Minimo 6 caracteres.")
                    else:
                        _db_manager.actualizar_password_usuario(user["id"], nueva)
                        st.success("Contrasena actualizada.")
            with c_del:
                if user.get("usuario") == (usuario_actual() or {}).get("usuario"):
                    st.caption("Tu usuario")
                elif st.button("Eliminar", key=f"usr_del_{user['id']}", use_container_width=True):
                    _db_manager.eliminar_usuario(user["id"])
                    st.rerun()

# Estilos visuales cargados desde assets/styles.css
# ==============================================================================
# 1. CARGA DE DATOS
# ==============================================================================
@st.cache_data(ttl=300)
def cargar_personal():
    nombre_archivo = "alumnos.csv"
    try:
        if os.path.exists(nombre_archivo):
            df = pd.read_csv(nombre_archivo, sep=";", encoding="utf-8")
            df.columns = df.columns.str.strip().str.upper()
            df['ORDEN_LIMP'] = pd.to_numeric(df['ORDEN'], errors='coerce')
            df['NOMBRE_COMPLETO'] = df['NOMBRE'].astype(str).str.strip().str.upper()
            df['DNI'] = df['DNI'].astype(str).str.strip()
            df['CE'] = df['CE'].astype(str).str.strip()
            df['GRADO'] = df['CURSO'].astype(str).str.strip().str.upper()
            df['AULA'] = df['AULA'].astype(str).str.strip()
            df = df.dropna(subset=['ORDEN_LIMP'])
            return df[['ORDEN_LIMP', 'AULA', 'GRADO', 'NOMBRE_COMPLETO', 'DNI', 'CE']].sort_values('ORDEN_LIMP')
        else:
            st.error("No se encontró 'alumnos.csv' en la carpeta del proyecto.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al cargar datos: {e}")
        return pd.DataFrame()

# ==============================================================================
# 2. INICIALIZACIÓN & DB
# ==============================================================================

# Inicializar base de datos
# Se inicializa solo cuando cambia la base/ruta. Evita lentitud en cada rerun.
firma_db_actual = getattr(_db_manager, "firma_base_datos", lambda: "default")()
if st.session_state.get("db_iniciada_firma") != firma_db_actual:
    init_db()
    st.session_state.db_iniciada_firma = firma_db_actual
    st.session_state.db_iniciada = True

aplicar_recuperacion_admin()

# Cargar personal antes del login para permitir formulario público protegido por token.
df = cargar_personal()
if df.empty:
    st.stop()

if formulario_plan_publico_solicitado():
    mostrar_formulario_plan_publico(df)
    st.stop()

if formulario_dinamico_publico_solicitado():
    mostrar_formulario_dinamico_publico(df)
    st.stop()

requerir_login()

info_bd = getattr(_db_manager, "info_base_datos", lambda: {"motor": "desconocido", "persistente": False, "ruta": ""})()
if not info_bd.get("persistente"):
    st.sidebar.warning(
        "⚠️ Base local SQLite: sirve para pruebas, pero puede vaciarse si el servidor se reinicia. "
        "Para guardar usuarios y datos permanentemente, configure DATABASE_URL en Secrets."
    )
    if info_bd.get("ruta"):
        st.sidebar.caption(f"BD local: {info_bd['ruta']}")
else:
    st.sidebar.success("✅ Base de datos persistente conectada")

if 'horarios_txt_importado' not in st.session_state:
    st.session_state.horarios_txt_importado = cargar_horarios_txt(r"C:\Users\admin\Desktop\horarios.txt")

TOTAL_ESCUADRON = len(df)
AULAS_UNICAS = sorted(df['AULA'].unique())

# Fecha del reporte
FECHA_STR = ahora_local().date().isoformat()
if 'fecha_reporte' not in st.session_state:
    fecha_guardada = obtener_config("fecha_reporte", None)
    fecha_reporte_inicial = ahora_local().date()
    if fecha_guardada:
        try:
            fecha_reporte_inicial = datetime.strptime(fecha_guardada, "%Y-%m-%d").date()
        except ValueError:
            fecha_reporte_inicial = ahora_local().date()
    if fecha_reporte_inicial < ahora_local().date():
        fecha_reporte_inicial = ahora_local().date()
        guardar_config("fecha_reporte", fecha_reporte_inicial.isoformat())
    st.session_state.fecha_reporte = fecha_reporte_inicial
FECHA_PARTE_STR = st.session_state.fecha_reporte.isoformat()

# Novedades
if 'novedades_lista' not in st.session_state:
    st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)

# Horarios config
if 'horarios_config' not in st.session_state:
    dia_reporte = DIAS_SEMANA[st.session_state.fecha_reporte.weekday()]
    db_hor = obtener_horarios_dia(dia_reporte) or obtener_horarios()
    st.session_state.horarios_config = {}
    for aula in AULAS_UNICAS:
        st.session_state.horarios_config[aula] = db_hor.get(normalizar_aula(aula), db_hor.get(aula, {
            "ent_m": "06:00", "sal_m": "12:00", "ent_t": "13:00", "sal_t": "19:00"
        }))

# Estado de aulas sincronizado con horarios del reporte
if 'estado_aulas' not in st.session_state or st.session_state.get('estado_aulas_fecha') != FECHA_PARTE_STR:
    db_estado = obtener_estado_aulas(FECHA_PARTE_STR)
    st.session_state.estado_aulas = {}
    st.session_state.estado_aulas_fecha = FECHA_PARTE_STR
    for aula in AULAS_UNICAS:
        aula_data = db_estado.get(aula, {})
        st.session_state.estado_aulas[aula] = {
            "estado_m": aula_data.get("estado_m", "EN INSTITUTO"),
            "estado_t": aula_data.get("estado_t", "EN INSTITUTO"),
            "salida_m": aula_data.get("salida_m"),
            "salida_t": aula_data.get("salida_t"),
            "ubicacion_m": aula_data.get("ubicacion_m", "EN AULA"),
            "ubicacion_t": aula_data.get("ubicacion_t", "EN AULA")
        }
sincronizar_ubicacion_con_horarios()

# Lista de almuerzo
if 'lista_almuerzo' not in st.session_state:
    st.session_state.lista_almuerzo = obtener_almuerzo(FECHA_PARTE_STR)

# Asistencia diaria
if 'estado_asistencia' not in st.session_state:
    st.session_state.estado_asistencia = obtener_asistencia(FECHA_PARTE_STR)

# Variables de control UI (¡ESTAS SON LAS QUE FALTABAN!)
if 'editando_idx' not in st.session_state:
    st.session_state.editando_idx = None

if 'sel_nov' not in st.session_state:
    st.session_state.sel_nov = None

def log_movimiento(modulo, accion, orden=None, nombre=None, aula=None, detalle=""):
    registrar_movimiento(FECHA_PARTE_STR, modulo, accion, orden, nombre, aula, detalle)


def estado_asistencia_por_ambito(ambito):
    if ambito == "AUSENTE":
        return "AUSENTE"
    if ambito == "INSTITUTO":
        return "PRESENTE EN INSTITUTO"
    return "PRESENTE EN ESCUADR?N"

def parsear_detalle_movimiento(detalle):
    datos = {"Motivo": "", "Presencia": "", "Desde": "", "Hasta": "", "Observación": detalle or ""}
    partes = [p.strip() for p in str(detalle or "").split("|")]
    if not partes:
        return datos

    datos["Motivo"] = partes[0] if partes else ""
    if len(partes) >= 2:
        if " a " in partes[1]:
            datos["Desde"], datos["Hasta"] = [p.strip() for p in partes[1].split(" a ", 1)]
            if len(partes) >= 3:
                datos["Observación"] = partes[2]
        else:
            datos["Presencia"] = partes[1]
            if len(partes) >= 3 and " a " in partes[2]:
                datos["Desde"], datos["Hasta"] = [p.strip() for p in partes[2].split(" a ", 1)]
                if len(partes) >= 4:
                    datos["Observación"] = partes[3]
            elif len(partes) >= 3:
                datos["Observación"] = partes[2]
    return datos

def limpiar_form_novedad():
    for key in ("sel_estado", "sel_ambito", "txt_detalle"):
        if key in st.session_state:
            del st.session_state[key]

if st.session_state.pop("limpiar_form_novedad_pendiente", False):
    limpiar_form_novedad()

# ==============================================================================
# 3. MÉTRICAS EN TIEMPO REAL (CON SINCRONIZACIÓN AUTOMÁTICA)
# ==============================================================================

# # 🔹 1. RECARGAR DATOS DESDE DB (Prioridad a session_state)
st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
st.session_state.lista_almuerzo = obtener_almuerzo(FECHA_PARTE_STR)

# Asistencia: NO sobrescribir si ya hay datos (prioridad a cambios manuales)
if not st.session_state.estado_asistencia:
    st.session_state.estado_asistencia = obtener_asistencia(FECHA_PARTE_STR)
else:
    # Sincronizar solo registros nuevos de DB
    db_asistencia = obtener_asistencia(FECHA_PARTE_STR)
    for orden, estado in db_asistencia.items():
        if orden not in st.session_state.estado_asistencia:
            st.session_state.estado_asistencia[orden] = estado

# 🔹 2. CÁLCULO DE MÉTRICAS (Se recalcula en cada interacción)
ambito_por_orden = {
    n['orden']: ambito_efectivo(n)
    for n in st.session_state.novedades_lista
}
ausentes_fijos = {orden for orden, ambito in ambito_por_orden.items() if ambito == "AUSENTE"}
presentes_instituto_novedad = {orden for orden, ambito in ambito_por_orden.items() if ambito in {"INSTITUTO", "ESCUADRON"}}
presentes_escuadron_novedad = {orden for orden, ambito in ambito_por_orden.items() if ambito == "ESCUADRON"}
ausentes_manuales = {orden for orden, estado in st.session_state.estado_asistencia.items() if estado == "AUSENTE"}
presentes_instituto_manuales = {
    orden for orden, estado in st.session_state.estado_asistencia.items()
    if estado in {"PRESENTE", "PRESENTE EN INSTITUTO", "PRESENTE EN ESCUADRÓN"}
}
presentes_escuadron_manuales = {
    orden for orden, estado in st.session_state.estado_asistencia.items()
    if estado in {"PRESENTE", "PRESENTE EN ESCUADRÓN"}
}

total_ausentes = ausentes_fijos | (ausentes_manuales - presentes_instituto_manuales)

en_instituto = 0
fuera_por_aula = 0
presentes_escuadron = 0
for _, row in df.iterrows():
    orden = row['ORDEN_LIMP']
    aula = row['AULA']
    if orden in total_ausentes: continue
    en_aula_instituto = st.session_state.estado_aulas.get(aula, {}).get('estado_m', 'EN INSTITUTO') == 'EN INSTITUTO'
    if orden in presentes_instituto_novedad or orden in presentes_instituto_manuales or en_aula_instituto:
        en_instituto += 1
        if orden in presentes_escuadron_novedad or orden in presentes_escuadron_manuales or (orden not in ambito_por_orden and orden not in presentes_instituto_manuales and en_aula_instituto):
            presentes_escuadron += 1
    else:
        fuera_por_aula += 1

disponibles = TOTAL_ESCUADRON - len(total_ausentes)
total_fuera = fuera_por_aula + len(total_ausentes)

df_presentes_escuadron_base = df[
    (~df['ORDEN_LIMP'].isin(total_ausentes)) &
    (
        df['ORDEN_LIMP'].isin(presentes_escuadron_novedad) |
        df['ORDEN_LIMP'].isin(presentes_escuadron_manuales) |
        ((~df['ORDEN_LIMP'].isin(ambito_por_orden.keys())) &
         (~df['ORDEN_LIMP'].isin(presentes_instituto_manuales)) &
         (df['AULA'].map(lambda aula: st.session_state.estado_aulas.get(aula, {}).get('estado_m', 'EN INSTITUTO')) == 'EN INSTITUTO'))
    )
]
df_formacion_0600_base = df[
    (~df['ORDEN_LIMP'].isin(total_ausentes)) &
    (
        df['ORDEN_LIMP'].isin(presentes_escuadron_novedad) |
        df['ORDEN_LIMP'].isin(presentes_escuadron_manuales) |
        ((~df['ORDEN_LIMP'].isin(ambito_por_orden.keys())) &
         (~df['ORDEN_LIMP'].isin(presentes_instituto_manuales)))
    )
]
df_presentes_primera = df_formacion_0600_base[
    df_formacion_0600_base['AULA'].map(aula_ingresa_primera_obligacion)
]
primera_total = len(df_presentes_primera)
primera_tercer_anio = len(df_presentes_primera[df_presentes_primera['GRADO'].map(es_tercer_anio)])
primera_aop = len(df_presentes_primera[df_presentes_primera['GRADO'].map(es_aop)])

ubicacion_dist = {"EN AULA": [], "URF": [], "EDUCACION FISICA": [], "EN INSTITUTO": [], "FORMACION": []}
for aula in AULAS_UNICAS:
    cfg = st.session_state.estado_aulas[aula]
    if cfg['estado_m'] == 'EN INSTITUTO':
        ubic = cfg.get('ubicacion_m', 'EN AULA')
        if ubic == "FORMACION":
            ubic = "EN INSTITUTO"
        ubic_norm = str(ubic).upper().replace(chr(205), "I").replace(chr(211), "O").replace("?", "I")
        if "FISICA" in ubic_norm:
            ubic = "EDUCACION FISICA"
        if ubic in ubicacion_dist:
            ubicacion_dist[ubic].append(len(df[df['AULA'] == aula]))

# Panel institucional y métricas principales
en_aula_count = sum(ubicacion_dist.get('EN AULA', []))
urf_count = sum(ubicacion_dist.get('URF', []))
edfis_key = next((k for k in ubicacion_dist.keys() if 'FISICA' in k.upper() or 'FÍSICA' in k.upper()), 'EDUCACIÓN FÍSICA')
edfis_count = sum(ubicacion_dist.get(edfis_key, []))
activ_count = sum(ubicacion_dist.get('EN INSTITUTO', []))

kpis = [
    ("Total", TOTAL_ESCUADRON, "Dotación registrada", "neutral"),
    ("Presentes", disponibles, f"Ausentes {len(total_ausentes)}", "ok" if disponibles == TOTAL_ESCUADRON else "warn"),
    ("En instituto", en_instituto, "Personal dentro", "ok"),
    ("En escuadrón", presentes_escuadron, "Sin comisión", "ok"),
    ("Fuera", total_fuera, "Ausentes o retirados", "alert" if total_fuera else "neutral"),
    ("1ra oblig.", primera_total, "Formados 06:00", "ok"),
    ("3er año", primera_tercer_anio, "En formación", "neutral"),
    ("AOP", primera_aop, "En formación", "neutral"),
    ("En aula", en_aula_count, f"{len(ubicacion_dist.get('EN AULA', []))} aulas", "neutral"),
    ("URF", urf_count, f"{len(ubicacion_dist.get('URF', []))} aulas", "neutral"),
    ("Ed. física", edfis_count, f"{len(ubicacion_dist.get(edfis_key, []))} aulas", "neutral"),
    ("Actividad", activ_count, f"{len(ubicacion_dist.get('EN INSTITUTO', []))} aulas", "neutral"),
]

kpi_html = "".join(
    f'<div class="rrhh-kpi rrhh-kpi-{status}"><span>{label}</span><strong>{value}</strong><small>{caption}</small></div>'
    for label, value, caption, status in kpis
)

def texto_obs_motivos(novedades):
    if not novedades:
        return "Sin novedades"
    motivos = pd.Series([n["estado"] for n in novedades]).value_counts().sort_index()
    partes = []
    for motivo, cantidad in motivos.items():
        label = str(motivo).lower()
        partes.append(f"{int(cantidad)} {label}")
    return ", ".join(partes)

def tarjeta_monitor_novedades(titulo, novedades):
    total = len(novedades)
    obs = texto_obs_motivos(novedades)
    return (
        '<div class="nov-card">'
        f'<div><span>{escape(titulo)}</span><strong>{total}</strong></div>'
        f'<p><b>OBS:</b> {escape(obs)}</p>'
        '</div>'
    )

novedades_ausentes = [n for n in st.session_state.novedades_lista if ambito_efectivo(n) == "AUSENTE"]
novedades_ausentes_tercero = [n for n in novedades_ausentes if es_tercer_anio(n.get("grado", ""))]
novedades_ausentes_aop = [n for n in novedades_ausentes if es_aop(n.get("grado", ""))]

monitor_items_html = (
    tarjeta_monitor_novedades("Ausentes de 3er año", novedades_ausentes_tercero)
    + tarjeta_monitor_novedades("Ausentes AOP", novedades_ausentes_aop)
)

def mostrar_monitor_novedades():
    st.markdown(f"""
    <div class="nov-monitor">
        <div class="nov-monitor-head">
            <div class="nov-monitor-title">Monitor de novedades</div>
            <div class="nov-monitor-sub">{len(st.session_state.novedades_lista)} activa(s)</div>
        </div>
        <div class="nov-monitor-grid">{monitor_items_html}</div>
    </div>
    """, unsafe_allow_html=True)

logo_uri = asset_data_uri(APP_LOGO_FILE)
logo_header_html = f'<img class="rrhh-logo" src="{logo_uri}" alt="Escuadrón H">' if logo_uri else '<div class="rrhh-emblem">EH</div>'
malvinas_flag_html = """
<span class="malvinas-flag" title="Islas Malvinas" aria-label="Islas Malvinas">
    <svg viewBox="0 0 64 36" role="img" focusable="false" aria-hidden="true">
        <path class="malvinas-island" d="M7 17.8c4.2-5.1 10.6-5.8 16.2-2.9 3.2 1.7 5.8 1.8 8.9.9 2.3-.7 4.7.8 4.2 3.1-.5 2.4-4 2.2-6.4 3.2-2.8 1.1-5.6 3.8-9.4 3.2-4.1-.7-5-4.2-8.7-4.2-2.8 0-6.7.8-4.8-3.3Z"/>
        <path class="malvinas-band" d="M10.3 18.4c7.6-1.2 14.4 1.8 23.2-.4.5.6.6 1.3.1 2-8.6 2.3-15.6-.7-23.9.5-.4-.7-.2-1.5.6-2.1Z"/>
        <path class="malvinas-island" d="M38.6 14.6c4.9-3.2 11.2-1.7 15.5 1.8 2.5 2 5.8 2.2 5.8 4.6 0 2.7-4.5 1.4-7 2.8-3.7 2-7.9 4.1-12 2.4-3.3-1.4-2.2-5.2-5.5-6.3-2.8-.9.7-3.6 3.2-5.3Z"/>
        <path class="malvinas-band" d="M40.2 17.7c5.8.5 10.3 3 16.5 2.8.4.7.2 1.4-.4 1.9-6.3.2-11.1-2.3-17.7-2.8.1-.8.6-1.4 1.6-1.9Z"/>
        <circle class="malvinas-sun" cx="32" cy="18" r="2.7"/>
    </svg>
</span>
"""
clock_inicial = ahora_local().strftime("%d/%m/%Y %H:%M:%S")

st.markdown(f"""
<section class="rrhh-panel">
    <div class="rrhh-head">
        <div class="rrhh-brand">
            {logo_header_html}
            <div>
                <p class="rrhh-eyebrow">Sistema compartido de control de personal</p>
                <h1 class="rrhh-title"><span class="rrhh-title-full">Escuadrón H "Cabo Marcelo Godoy"</span><span class="rrhh-title-short">Escuadron H</span></h1>
                <p class="rrhh-subtitle">Parte diario, novedades, presentismo, ubicación y reportes en tiempo real.</p>
            </div>
        </div>
        <div class="rrhh-status-box">
            <div class="rrhh-date rrhh-clock">{malvinas_flag_html}<span id="ba-clock">{clock_inicial} Hs</span></div>
            <div class="rrhh-date">Parte: {st.session_state.fecha_reporte.strftime('%d/%m/%Y')}</div>
        </div>
    </div>
    <div class="rrhh-access">
        <div class="rrhh-access-card"><span>Trabajo multiusuario</span><strong>Control centralizado para guardia y conducción</strong></div>
        <div class="rrhh-access-card"><span>Seguimiento activo</span><strong>Novedades activas con seguimiento automático</strong></div>
        <div class="rrhh-access-card"><span>Reportes</span><strong>Exportación institucional lista para compartir</strong></div>
    </div>
    <div class="rrhh-kpi-grid">{kpi_html}</div>
</section>
""", unsafe_allow_html=True)

components.html(
    """
    <script>
    (function () {
        function updateBuenosAiresClock() {
            const el = window.parent.document.getElementById("ba-clock");
            if (!el) return;
            const now = new Date();
            const date = new Intl.DateTimeFormat("es-AR", {
                timeZone: "America/Argentina/Buenos_Aires",
                day: "2-digit",
                month: "2-digit",
                year: "numeric"
            }).format(now);
            const time = new Intl.DateTimeFormat("es-AR", {
                timeZone: "America/Argentina/Buenos_Aires",
                hour: "2-digit",
                minute: "2-digit",
                second: "2-digit",
                hour12: false
            }).format(now);
            el.textContent = `${date} ${time} Hs`;
        }
        updateBuenosAiresClock();
        window.setInterval(updateBuenosAiresClock, 1000);
    })();
    </script>
    """,
    height=0,
)

st.divider()
with st.expander("⚙️ Acciones rápidas y mantenimiento", expanded=False):
    st.caption("Usá estas opciones solo cuando necesites refrescar datos o reiniciar la asistencia del día.")
    col_sync, col_reset = st.columns(2)
    with col_sync:
        if st.button("🔄 Sincronizar datos", key="sync_btn", help="Fuerza la recarga desde base de datos", use_container_width=True):
            st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
            st.session_state.estado_asistencia = obtener_asistencia(FECHA_PARTE_STR)
            st.session_state.lista_almuerzo = obtener_almuerzo(FECHA_PARTE_STR)
            st.success("✅ Datos sincronizados correctamente")
            st.rerun()
    with col_reset:
        if st.button("🚨 Reiniciar asistencia", key="reset_asistencia", help="Pone a TODOS en PRESENTE", use_container_width=True):
            with _db_manager.get_db() as conn:
                _db_manager.run(conn, "DELETE FROM asistencia_diaria WHERE fecha=?", (FECHA_PARTE_STR,))
                conn.commit()
            st.session_state.estado_asistencia = {}
            st.success("✅ Asistencia reiniciada. Todos en PRESENTE.")
            st.rerun()
# ==============================================================================
# 4. PESTAÑAS
# ==============================================================================
tab_labels = ["📅 Día y horarios", "📝 Novedades", "📍 Ubicación", "🍽️ Racionamiento", "📇 Legajos y contactos", "📊 Reportes"]
if es_admin():
    tab_labels.append("👥 Usuarios")
tabs_creadas = st.tabs(tab_labels)
tab_config, tab_nov, tab_seg, tab_alm, tab_plan, tab_res = tabs_creadas[:6]
tab_usuarios = tabs_creadas[6] if es_admin() else None

# --- TAB: CONFIGURACIÓN ---
with tab_config:
    st.subheader("📅 Configuración del día y horarios")
    fecha_actual_reporte = st.session_state.fecha_reporte
    fecha_reporte_nueva = st.date_input(
        "Fecha del Reporte",
        fecha_actual_reporte,
        help="Selecciona la fecha y presiona Guardar fecha del reporte para dejarla fija."
    )
    col_fecha_guardar, col_fecha_info = st.columns([1, 2])
    with col_fecha_guardar:
        guardar_fecha_btn = st.button("Guardar fecha del reporte", type="primary", use_container_width=True)
    with col_fecha_info:
        st.caption(f"Fecha guardada actual: {fecha_actual_reporte.strftime('%d/%m/%Y')}")
    if guardar_fecha_btn:
        st.session_state.fecha_reporte = fecha_reporte_nueva
        guardar_config("fecha_reporte", fecha_reporte_nueva.isoformat())
        db_hor = obtener_horarios_dia(DIAS_SEMANA[fecha_reporte_nueva.weekday()]) or obtener_horarios()
        st.session_state.horarios_config = {}
        for aula in AULAS_UNICAS:
            st.session_state.horarios_config[aula] = db_hor.get(normalizar_aula(aula), db_hor.get(aula, {
                "ent_m": "06:00", "sal_m": "12:00", "ent_t": "13:00", "sal_t": "19:00"
            }))
        st.session_state.pop('estado_aulas', None)
        st.session_state.pop('estado_aulas_fecha', None)
        st.success("Fecha del reporte guardada")
        st.rerun()
    dia_reporte = DIAS_SEMANA[st.session_state.fecha_reporte.weekday()]
    st.caption(f"Horarios cargados para: {dia_reporte}. Podes editarlos y guardarlos para ese dia.")
    st.divider()
    for aula in AULAS_UNICAS:
        cfg = st.session_state.horarios_config[aula]
        with st.expander(f"**{aula}**", expanded=False):
            c1, c2, c3, c4 = st.columns(4)
            cfg['ent_m'] = c1.text_input("Entrada Mañana", cfg['ent_m'], key=f"em_{aula}")
            cfg['sal_m'] = c2.text_input("Salida Mañana", cfg['sal_m'], key=f"sm_{aula}")
            cfg['ent_t'] = c3.text_input("Entrada Tarde", cfg['ent_t'], key=f"et_{aula}")
            cfg['sal_t'] = c4.text_input("Salida Tarde", cfg['sal_t'], key=f"st_{aula}")
    
    if st.button("💾 Guardar configuración de horarios", type="primary"):
        for aula in AULAS_UNICAS:
            guardar_horarios_dia(normalizar_aula(aula), dia_reporte, st.session_state.horarios_config[aula])
            guardar_horarios(normalizar_aula(aula), st.session_state.horarios_config[aula])
        st.success("✅ Horarios guardados en base de datos")
        st.rerun()

if tab_usuarios is not None:
    with tab_usuarios:
        panel_admin_usuarios()

# --- TAB: NOVEDADES ---
with tab_nov:
    st.divider()
    edit_idx = st.session_state.editando_idx
    es_edicion = edit_idx is not None

    # Validación de seguridad por si la lista cambió mientras editabas
    if es_edicion and not (0 <= edit_idx < len(st.session_state.novedades_lista)):
        st.session_state.editando_idx = None
        st.warning("⚠️ La novedad que editabas cambió. Se reinició el formulario.")
        st.rerun()

    st.subheader("✏️ Editando novedad" if es_edicion else "➕ Registrar novedad")

    data = None
    if es_edicion:
        nov = st.session_state.novedades_lista[edit_idx]
        st.info(f"Editando a: **{nov['nombre']}**")
        data = nov
        
        st.divider()
        st.markdown(f"### Control de presencia: **{data['nombre']}**")
        
        orden = data["orden"]
        nombre_asp = data["nombre"]

        col_btn_aus, col_btn_p, col_btn_a = st.columns(3)
        
        with col_btn_aus:
            if st.button("Ausente", type="secondary", use_container_width=True, key="btn_aus_edit"):
                st.session_state.sel_ambito = "AUSENTE"
                st.toast(f"{nombre_asp} preparado como ausente")
                st.rerun()

        with col_btn_p:
            if st.button("Presente en instituto", type="secondary", use_container_width=True, key="btn_pres_inst_edit"):
                st.session_state.sel_ambito = "INSTITUTO"
                st.toast(f"{nombre_asp} preparado como presente en instituto")
                st.rerun()
                
        with col_btn_a:
            if st.button("Presente en escuadrón", type="primary", use_container_width=True, key="btn_pres_esc_edit"):
                st.session_state.sel_ambito = "ESCUADRON"
                st.toast(f"{nombre_asp} preparado como presente en escuadrón")
                st.rerun()
        st.divider()

    else:
        # Lógica de búsqueda y selección (Modo Registro)
        search = live_search_input("Buscar aspirante:", "Nombre, DNI o CE", "search_nov")
        if search.strip() and not st.session_state.sel_nov:
            s = search.strip().upper()
            res = df[
                (df['NOMBRE_COMPLETO'].str.contains(s, na=False)) |
                (df['DNI'].str.contains(s, na=False)) |
                (df['CE'].str.contains(s, na=False))
            ]
            if not res.empty:
                st.markdown("### Resultados de búsqueda")
                for i, (_, r) in enumerate(res.head(5).iterrows()):
                    c1, c2 = st.columns([4, 1])
                    with c1: st.markdown(f"**{r['NOMBRE_COMPLETO']}** | DNI: {r['DNI']} | CE: {r['CE']}")
                    with c2:
                        if st.button("👆 Seleccionar", key=f"sel_{i}"):
                            limpiar_form_novedad()
                            st.session_state.sel_nov = r.to_dict()
                            st.session_state.search_nov = ""
                            st.rerun()

        if st.session_state.sel_nov:
            data = st.session_state.sel_nov
            orden = data["ORDEN_LIMP"]
            nombre_asp = data.get('NOMBRE_COMPLETO', data.get('nombre', 'Aspirante'))
            st.divider()
            st.markdown(f"### Control de presencia: **{nombre_asp}**")

            col_btn_aus, col_btn_p, col_btn_a, col_btn_c = st.columns([2, 2, 2, 1])

            with col_btn_aus:
                if st.button("Ausente", type="secondary", use_container_width=True, key="btn_aus"):
                    st.session_state.sel_ambito = "AUSENTE"
                    st.toast(f"{nombre_asp} preparado como ausente")
                    st.rerun()

            with col_btn_p:
                if st.button("Presente en instituto", type="secondary", use_container_width=True, key="btn_pres_inst"):
                    st.session_state.sel_ambito = "INSTITUTO"
                    st.toast(f"{nombre_asp} preparado como presente en instituto")
                    st.rerun()

            with col_btn_a:
                if st.button("Presente en escuadrón", type="primary", use_container_width=True, key="btn_pres_esc"):
                    st.session_state.sel_ambito = "ESCUADRON"
                    st.toast(f"{nombre_asp} preparado como presente en escuadrón")
                    st.rerun()

            with col_btn_c:
                if st.button("Cambiar", use_container_width=True, help="Cambiar aspirante", key="btn_clear_sel"):
                    limpiar_form_novedad()
                    st.session_state.sel_nov = None
                    st.session_state.search_nov = ""
                    st.rerun()
            st.divider()

    # 🔹 FORMULARIO DE NOVEDAD (Visible si hay data, tanto en edición como en registro)
    if data is not None:
        st.markdown("### ⚙️ Detalles de Novedad")
        c1, c2 = st.columns(2)
        with c1:
            opts = [
    "AUSENTE",
    "ART",
    "DAF",
    "LES",
    "LAO",
    "SSD",
    "COMISIÓN",
    "AUTORIZADO",
    "ENTRANTE GUARDIA DIURNA",
    "ENTRANTE GUARDIA NOCTURNA",
    "DESCANSO DE GUARDIA"
]
            current_estado = data.get('estado', "ART")
            if current_estado not in opts:
                current_estado = "ART"
            if st.session_state.get("sel_estado") not in opts:
                st.session_state.sel_estado = current_estado
            idx_opts = opts.index(current_estado) if current_estado in opts else 0
            est = st.selectbox("Situación:", opts, index=idx_opts, key="sel_estado")
        with c2:
            det = st.text_input("Detalle:", value=data.get('detalle', ''), key="txt_detalle")

        ambito_actual = st.session_state.get("sel_ambito") or data.get('ambito') or ambito_por_defecto(est)
        ambito_keys = list(AMBITOS_NOVEDAD.keys())
        ambito_idx = ambito_keys.index(ambito_actual) if ambito_actual in ambito_keys else 0
        ambito = st.radio(
            "Presencia real:",
            ambito_keys,
            index=ambito_idx,
            format_func=lambda x: AMBITOS_NOVEDAD[x],
            key="sel_ambito",
            horizontal=True
        )

        fecha_form_default = st.session_state.fecha_reporte
        cf1, cf2 = st.columns(2)
        with cf1:
            if es_edicion:
                try:
                    fi_val = datetime.strptime(data['fecha_ini'], '%d%b%y').date()
                except Exception:
                    fi_val = fecha_form_default
                fi = st.date_input("Desde:", value=fi_val, key="date_ini").strftime('%d%b%y').upper()
            else:
                fi = st.date_input("Desde:", value=fecha_form_default, key="date_ini2").strftime('%d%b%y').upper()
        with cf2:
            is_no = (data.get('fecha_fin') == "N/O") if es_edicion else False
            sin_fin = st.checkbox("Sin término", value=is_no, key="chk_sintermino")
            if sin_fin:
                ff = "N/O"
            else:
                if es_edicion and not is_no:
                    try:
                        ff_val = datetime.strptime(data['fecha_fin'], '%d%b%y').date()
                    except Exception:
                        ff_val = fecha_form_default
                    ff = st.date_input("Hasta:", value=ff_val, key="date_fin").strftime('%d%b%y').upper()
                else:
                    ff = st.date_input("Hasta:", value=fecha_form_default, key="date_fin2").strftime('%d%b%y').upper()

        b1, b2 = st.columns([3, 1])
        with b1:
            if es_edicion:
                if st.button("💾 Guardar Cambios", type="primary", use_container_width=True, key="btn_save_edit"):
                    nov_id = st.session_state.novedades_lista[edit_idx]['id']
                    actualizar_novedad(nov_id, {"estado": est, "detalle": det.upper(), "fecha_ini": fi, "fecha_fin": ff, "ambito": ambito})
                    actualizar_asistencia(FECHA_PARTE_STR, data.get("orden"), estado_asistencia_por_ambito(ambito))
                    st.session_state.estado_asistencia[data.get("orden")] = estado_asistencia_por_ambito(ambito)
                    log_movimiento("Novedades", "EDITAR NOVEDAD", data.get("orden"), data.get("nombre"), data.get("aula"), f"{est} | {AMBITOS_NOVEDAD.get(ambito, ambito)} | {fi} a {ff} | {det.upper()}")
                    st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
                    st.session_state.editando_idx = None
                    st.session_state.limpiar_form_novedad_pendiente = True
                    st.success("✅ Novedad y asistencia actualizadas")
                    st.rerun()
            else:
                if st.button("💾 Grabar Novedad", use_container_width=True, key="btn_save_new"):
                    nombre_asp = data.get('NOMBRE_COMPLETO', data.get('nombre', 'Aspirante'))
                    agregar_novedad({
                        "orden": int(data["ORDEN_LIMP"]), "grado": data["GRADO"],
                        "nombre": nombre_asp, "dni": data["DNI"], "ce": data["CE"],
                        "aula": data["AULA"], "estado": est, "detalle": det.upper(),
                        "fecha_ini": fi, "fecha_fin": ff, "ambito": ambito
                    })
                    actualizar_asistencia(FECHA_PARTE_STR, int(data["ORDEN_LIMP"]), estado_asistencia_por_ambito(ambito))
                    st.session_state.estado_asistencia[int(data["ORDEN_LIMP"])] = estado_asistencia_por_ambito(ambito)
                    log_movimiento("Novedades", "ALTA NOVEDAD", int(data["ORDEN_LIMP"]), nombre_asp, data["AULA"], f"{est} | {AMBITOS_NOVEDAD.get(ambito, ambito)} | {fi} a {ff} | {det.upper()}")
                    st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
                    st.session_state.sel_nov = None
                    st.session_state.limpiar_form_novedad_pendiente = True
                    st.success(f"✅ Novedad grabada para {nombre_asp}")
                    st.rerun()
        with b2:
            if st.button("🚫 Cancelar", use_container_width=True, key="btn_cancel"):
                st.session_state.limpiar_form_novedad_pendiente = True
                st.session_state.editando_idx = None
                st.session_state.sel_nov = None
                st.rerun()
    else:
        if not es_edicion:
            st.info("🔍 Busca un aspirante para registrar novedad o asistencia.")

    st.divider()
    st.subheader("Novedades activas")
    if st.session_state.novedades_lista:
        st.caption("Panel editable para modificar motivo, presencia real, fechas o cerrar una novedad antes de su vencimiento.")
        for idx, nov in enumerate(st.session_state.novedades_lista):
            ambito_lbl = AMBITOS_NOVEDAD.get(ambito_efectivo(nov), "Sin definir")
            with st.container(border=True):
                c_info, c_edit, c_del = st.columns([6, 1, 1])
                with c_info:
                    st.markdown(f"**{nov['nombre']}** | **{nov['estado']}**")
                    st.caption(f"{nov['fecha_ini']} a {nov['fecha_fin']} | {ambito_lbl} | Aula {nov.get('aula', '-')} | DNI {nov.get('dni', '-')} | CE {nov.get('ce', '-')}")
                    if nov.get('detalle'):
                        st.caption(f"Detalle: {nov['detalle']}")
                with c_edit:
                    if st.button("Editar", key=f"edit_nov_activa_{idx}", use_container_width=True):
                        limpiar_form_novedad()
                        st.session_state.editando_idx = idx
                        st.session_state.sel_nov = None
                        st.rerun()
                with c_del:
                    if st.button("Quitar", key=f"del_nov_activa_{idx}", use_container_width=True):
                        log_movimiento("Novedades", "ELIMINAR NOVEDAD", nov.get("orden"), nov.get("nombre"), nov.get("aula"), f"{nov.get('estado')} | {nov.get('fecha_ini')} a {nov.get('fecha_fin')} | {nov.get('detalle')}")
                        eliminar_novedad(nov['id'])
                        st.session_state.estado_asistencia[nov['orden']] = "PRESENTE"
                        actualizar_asistencia(FECHA_PARTE_STR, nov['orden'], "PRESENTE")
                        st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
                        st.toast("Novedad eliminada y asistencia actualizada")
                        st.rerun()
    else:
        st.info("No hay novedades activas para la fecha seleccionada.")

# --- TAB: SEGUIMIENTO ---
with tab_seg:
    st.subheader("Ubicación del personal por aula")
    
    # Selector de turno
    turno_act = st.radio("Seleccionar Turno:", ["🌅 MAÑANA", "🌆 TARDE"], horizontal=True, label_visibility="collapsed")
    prefijo = "m" if turno_act == "🌅 MAÑANA" else "t"

    st.caption("Resumen compacto por aula. Abrí solo el aula que necesites modificar.")
    locations = [("Aula", "EN AULA"), ("URF", "URF"), ("Ed. física", "EDUCACIÓN FÍSICA"), ("Actividad", "EN INSTITUTO")]
    grid_cols = st.columns(2)

    for idx, aula in enumerate(AULAS_UNICAS):
        with grid_cols[idx % 2]:
            cfg = st.session_state.estado_aulas[aula]
            estado_key = f"estado_{prefijo}"
            salida_key = f"salida_{prefijo}"
            ubic_key = f"ubicacion_{prefijo}"
            alumnos = df[df['AULA'] == aula]
            total = len(alumnos)
            ausentes = sum(1 for n in st.session_state.novedades_lista if n['aula'] == aula and ambito_efectivo(n) == 'AUSENTE')
            presentes = total - ausentes
            is_inside = cfg[estado_key] == 'EN INSTITUTO'
            ubicacion_actual = cfg.get(ubic_key, 'EN AULA')
            estado_label = ubicacion_actual if is_inside else "FUERA"
            estado_color = "#22C55E" if is_inside else "#EF4444"
            salida_txt = f" | Salida: {cfg[salida_key]}" if cfg[salida_key] else ""

            st.markdown(
                f"""
                <div style="border:1px solid rgba(148,163,184,.22); border-radius:8px; padding:.65rem .75rem; margin:.35rem 0; background:rgba(255,255,255,.035);">
                    <div style="display:flex; justify-content:space-between; gap:.5rem; align-items:center;">
                        <strong style="font-size:1rem;">{aula}</strong>
                        <span style="font-size:.78rem; color:{estado_color}; font-weight:700;">{estado_label}</span>
                    </div>
                    <div style="font-size:.8rem; color:#A7B0BE; margin-top:.25rem;">Total {total} | Presentes {presentes} | Ausentes {ausentes}{salida_txt}</div>
                </div>
                """,
                unsafe_allow_html=True
            )

            with st.expander(f"Gestionar {aula}", expanded=False):
                if is_inside and ubicacion_actual == "FORMACION":
                    st.info("Presentacion/formacion en instituto. Las actividades manuales se habilitan desde las 07:00.")
                elif is_inside:
                    loc_cols = st.columns(4)
                    for i, (label, value) in enumerate(locations):
                        with loc_cols[i]:
                            is_active = ubicacion_actual == value
                            btn_type = "primary" if is_active else "secondary"
                            if st.button(label, key=f"loc_{prefijo}_{aula}_{i}", type=btn_type, use_container_width=True):
                                st.session_state.estado_aulas[aula][ubic_key] = value
                                guardar_estado_aula(FECHA_PARTE_STR, aula, cfg['estado_m'], cfg['estado_t'], cfg.get('salida_m'), cfg.get('salida_t'), cfg.get('ubicacion_m', 'EN AULA'), cfg.get('ubicacion_t', 'EN AULA'))
                                st.rerun()

                    if st.button("Retirar aula", key=f"out_{prefijo}_{aula}", use_container_width=True):
                        st.session_state.estado_aulas[aula][estado_key] = 'FUERA'
                        st.session_state.estado_aulas[aula][salida_key] = ahora_local().strftime("%H:%M")
                        guardar_estado_aula(FECHA_PARTE_STR, aula, cfg['estado_m'], cfg['estado_t'], cfg.get('salida_m'), cfg.get('salida_t'), cfg.get('ubicacion_m', 'EN AULA'), cfg.get('ubicacion_t', 'EN AULA'))
                        st.rerun()
                else:
                    if cfg.get(salida_key) == "FRANCO":
                        st.info("FRANCO por horario cargado en Dias y horarios. Se habilita automaticamente al llegar el ingreso.")
                    else:
                        st.warning("Aula fuera del instituto. Reingresar para habilitar ubicacion.")
                        if st.button("Reingresar aula", key=f"in_{prefijo}_{aula}", use_container_width=True):
                            st.session_state.estado_aulas[aula][estado_key] = 'EN INSTITUTO'
                            st.session_state.estado_aulas[aula][salida_key] = None
                            guardar_estado_aula(FECHA_PARTE_STR, aula, cfg['estado_m'], cfg['estado_t'], cfg.get('salida_m'), cfg.get('salida_t'), cfg.get('ubicacion_m', 'EN AULA'), cfg.get('ubicacion_t', 'EN AULA'))
                            st.rerun()

# --- TAB: ALMUERZO ---
with tab_alm:
    st.subheader("Control de Personal que Almuerza")
    search = live_search_input("Buscar aspirante:", "Nombre, DNI o CE", "search_alm")
    if search.strip():
        s = search.strip().upper()
        res = df[
            (df['NOMBRE_COMPLETO'].str.contains(s, na=False)) |
            (df['DNI'].str.contains(s, na=False)) |
            (df['CE'].str.contains(s, na=False))
        ]
        if not res.empty:
            st.markdown("### Resultados de búsqueda")
            for i, (_, r) in enumerate(res.head(10).iterrows()):
                c1, c2 = st.columns([4, 1])
                with c1: 
                    st.markdown(f"**{r['NOMBRE_COMPLETO']}** | {r['AULA']} | Orden: {r['ORDEN_LIMP']}")
                with c2:
                    if r['ORDEN_LIMP'] in st.session_state.lista_almuerzo:
                        st.button("✅ Marcado", key=f"marked_{i}", use_container_width=True, disabled=True)
                    else:
                        if st.button("➕ Marcar", key=f"m_{i}", use_container_width=True):
                            st.session_state.lista_almuerzo.add(r['ORDEN_LIMP'])
                            agregar_almuerzo(FECHA_PARTE_STR, r['ORDEN_LIMP'])
                            if hasattr(_db_manager, "registrar_almuerzo_historial"):
                                _db_manager.registrar_almuerzo_historial(FECHA_PARTE_STR, int(r['ORDEN_LIMP']), r['NOMBRE_COMPLETO'], r['AULA'])
                            st.rerun()
    
    st.divider()
    st.subheader("Lista de Almuerzo")
    if st.session_state.lista_almuerzo:
        df_lista = (
            df[df['ORDEN_LIMP'].isin(st.session_state.lista_almuerzo)]
            .sort_values('ORDEN_LIMP')
            .reset_index(drop=True)
        )
        tabla_almuerzo = pd.DataFrame({
            "Quitar": False,
            "Nro": range(1, len(df_lista) + 1),
            "Orden": df_lista["ORDEN_LIMP"].astype(int),
            "Nombre": df_lista["NOMBRE_COMPLETO"],
            "Grado": df_lista["GRADO"],
            "CE": df_lista["CE"],
            "DNI": df_lista["DNI"],
            "Aula": df_lista["AULA"],
        })
        st.caption(f"Total que almuerzan: {len(tabla_almuerzo)}")
        altura_tabla_almuerzo = min(330, 38 + (len(tabla_almuerzo) + 1) * 34)
        tabla_editada = st.data_editor(
            tabla_almuerzo,
            hide_index=True,
            use_container_width=True,
            height=altura_tabla_almuerzo,
            disabled=["Nro", "Orden", "Nombre", "Grado", "CE", "DNI", "Aula"],
            column_config={
                "Quitar": st.column_config.CheckboxColumn("Quitar", help="Marcar para quitar de la lista"),
                "Nro": st.column_config.NumberColumn("Nro", width="small"),
                "Orden": st.column_config.NumberColumn("Orden", width="small"),
                "Nombre": st.column_config.TextColumn("Nombre", width="large"),
                "Grado": st.column_config.TextColumn("Grado", width="medium"),
                "CE": st.column_config.TextColumn("CE", width="small"),
                "DNI": st.column_config.TextColumn("DNI", width="medium"),
                "Aula": st.column_config.TextColumn("Aula", width="small"),
            },
            key="tabla_almuerzo_compacta",
        )

        col_quitar, col_vaciar = st.columns([1, 1])
        with col_quitar:
            if st.button("Quitar seleccionados", use_container_width=True, key="remove_selected_lunch"):
                seleccionados = tabla_editada[tabla_editada["Quitar"]]
                if seleccionados.empty:
                    st.toast("No marcaste ningun aspirante para quitar")
                else:
                    for orden in seleccionados["Orden"].astype(int).tolist():
                        st.session_state.lista_almuerzo.discard(orden)
                        quitar_almuerzo(FECHA_PARTE_STR, orden)
                    st.toast(f"Se quitaron {len(seleccionados)} aspirante(s) de la lista")
                    st.rerun()
        with col_vaciar:
            if st.button("Vaciar lista completa", type="secondary", key="clear_all_lunch", use_container_width=True):
                for orden in list(st.session_state.lista_almuerzo):
                    quitar_almuerzo(FECHA_PARTE_STR, orden)
                st.session_state.lista_almuerzo.clear()
                st.rerun()
    else:
        st.info("Aun no hay personal marcado para almorzar.")

    st.divider()
    st.subheader("Historial de racionamiento")
    col_desde_rac, col_hasta_rac = st.columns(2)
    with col_desde_rac:
        fecha_desde_rac = st.date_input("Desde", datetime.now().date(), key="fecha_racionamiento_desde")
    with col_hasta_rac:
        fecha_hasta_rac = st.date_input("Hasta", datetime.now().date(), key="fecha_racionamiento_hasta")

    if fecha_desde_rac > fecha_hasta_rac:
        st.warning("La fecha desde no puede ser mayor que la fecha hasta.")
        hist_alm = []
    elif hasattr(_db_manager, "obtener_historial_almuerzo"):
        hist_alm = _db_manager.obtener_historial_almuerzo(fecha_desde_rac.isoformat(), fecha_hasta_rac.isoformat())
    else:
        hist_alm = []

    if hist_alm:
        df_hist_alm = pd.DataFrame(hist_alm)
        for col in ["nombre", "aula", "orden", "fecha", "fecha_hora"]:
            if col not in df_hist_alm.columns:
                df_hist_alm[col] = ""

        df_hist_alm["_orden_num"] = pd.to_numeric(df_hist_alm["orden"], errors="coerce")
        aspirantes_rac = (
            df_hist_alm.assign(_label=df_hist_alm.apply(
                lambda r: f"{int(r['_orden_num']) if pd.notna(r['_orden_num']) else ''} - {str(r['nombre']).strip()}".strip(" -"),
                axis=1,
            ))["_label"]
            .dropna()
            .drop_duplicates()
            .sort_values()
            .tolist()
        )
        aspirante_rac = st.selectbox(
            "Filtrar por aspirante",
            aspirantes_rac,
            index=None,
            placeholder="Escribi o deja vacio para ver todo",
            key="filtro_aspirante_racionamiento",
        )
        if aspirante_rac:
            orden_filtrado = aspirante_rac.split(" - ", 1)[0].strip()
            if orden_filtrado.isdigit():
                df_hist_alm = df_hist_alm[df_hist_alm["_orden_num"] == int(orden_filtrado)]

        st.caption(
            f"Periodo: {fecha_desde_rac.strftime('%d/%m/%Y')} al {fecha_hasta_rac.strftime('%d/%m/%Y')} | "
            f"Registros: {len(df_hist_alm)}"
        )

        if df_hist_alm.empty:
            st.info("No hay registros para el aspirante seleccionado en ese rango.")
        else:
            detalle_alm = df_hist_alm.rename(columns={
                "fecha_hora": "Fecha/hora registro",
                "fecha": "Fecha",
                "orden": "Orden",
                "nombre": "Nombre",
                "aula": "Aula",
            })
            detalle_alm = detalle_alm[["Fecha", "Fecha/hora registro", "Orden", "Nombre", "Aula"]]
            st.dataframe(detalle_alm, use_container_width=True, hide_index=True, height=300)

            resumen_alm = (
                df_hist_alm.groupby(["orden", "nombre", "aula"], dropna=False)
                .size()
                .reset_index(name="Veces en el rango")
                .sort_values(["Veces en el rango", "nombre"], ascending=[False, True])
                .rename(columns={"orden": "Orden", "nombre": "Nombre", "aula": "Aula"})
            )
            with st.expander("Resumen por aspirante", expanded=True):
                st.dataframe(resumen_alm, use_container_width=True, hide_index=True)
    else:
        st.info("No hay registros de racionamiento para el rango seleccionado.")

    st.divider()
    if st.session_state.lista_almuerzo:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RACIONAMIENTO"
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "PARTE DE RACIONAMIENTO - ESCUADRÓN H"
        ws['A1'].font = excel_font(bold=True, size=15, color=EXCEL_WHITE)
        ws['A1'].fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A2'] = f"Fecha: {st.session_state.fecha_reporte.strftime('%d/%m/%Y')} | Total: {len(st.session_state.lista_almuerzo)}"
        ws['A2'].font = excel_font(italic=True, size=10, color=EXCEL_TEXT_MUTED)
        ws.row_dimensions[1].height = 25
        
        headers = ["Nro", "NOMBRE COMPLETO", "GRADO", "CE", "DNI", "AULA"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = PatternFill(start_color=EXCEL_OLIVE, end_color=EXCEL_OLIVE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin", color=EXCEL_BORDER), right=Side(style="thin", color=EXCEL_BORDER), top=Side(style="thin", color=EXCEL_BORDER), bottom=Side(style="thin", color=EXCEL_BORDER))
            
        row = 5
        for nro, orden in enumerate(sorted(st.session_state.lista_almuerzo), 1):
            p = df[df['ORDEN_LIMP'] == orden].iloc[0]
            ws.cell(row=row, column=1, value=nro)
            ws.cell(row=row, column=2, value=p['NOMBRE_COMPLETO'])
            ws.cell(row=row, column=3, value=p['GRADO'])
            ws.cell(row=row, column=4, value=p['CE'])
            ws.cell(row=row, column=5, value=p['DNI'])
            ws.cell(row=row, column=6, value=p['AULA'])
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = Border(left=Side(style="thin", color=EXCEL_BORDER), right=Side(style="thin", color=EXCEL_BORDER), top=Side(style="thin", color=EXCEL_BORDER), bottom=Side(style="thin", color=EXCEL_BORDER))
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center" if c in [1,3,6] else "left")
            row += 1
            
        for col, w in zip("ABCDEF", [10, 35, 12, 12, 15, 12]): 
            ws.column_dimensions[col].width = w
        output = f"RACIONAMIENTO_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        st.download_button(
            "GENERAR PARTE DE RACIONAMIENTO",
            data=excel_bytes(wb),
            file_name=output,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            key="download_parte_racionamiento"
        )

# ==================== PLAN DE LLAMADA (CONTACTOS) ====================
# --- TAB: PLAN DE LLAMADA ---
with tab_plan:
    st.subheader("📞 Plan de Llamada - Base de Contactos")
    st.info("Registra domicilios y contactos de emergencia para cada personal del escuadrón.")

    with st.expander("📋 Formularios dinámicos", expanded=False):
        panel_formularios_dinamicos(df)

    with st.expander("🔗 Formulario público para datos de licencia", expanded=False):
        form_enabled = valor_verdadero(obtener_secret_o_env("PLAN_FORM_ENABLED", "false"))
        form_token = str(obtener_secret_o_env("PLAN_FORM_TOKEN", "")).strip()
        if form_enabled and form_token:
            st.success("Formulario público habilitado.")
            st.caption("Agregá este final a la URL de tu app y compartilo solo con el personal autorizado:")
            st.code(f"?form=plan_llamada&token={form_token}", language="text")
            st.caption("Ejemplo: https://TU-APP.streamlit.app/?form=plan_llamada&token=TU_TOKEN")
        else:
            st.warning("Formulario público deshabilitado. Para habilitarlo, agregá en Secrets:")
            st.code('PLAN_FORM_ENABLED = "true"\nPLAN_FORM_TOKEN = "CAMBIAR_TOKEN_SEGURO"', language="toml")
        st.caption("El formulario valida DNI/CE contra alumnos.csv, no modifica la base de personal y guarda solo los campos de licencia en plan_llamada.")
    
    search = live_search_input("Buscar personal:", "Nombre, DNI, CE o Orden", "search_plan")
    
    if search.strip():
        s = search.strip().upper()
        res = df[
            (df['NOMBRE_COMPLETO'].str.contains(s, na=False)) |
            (df['DNI'].str.contains(s, na=False)) |
            (df['CE'].str.contains(s, na=False)) |
            (df['ORDEN_LIMP'].astype(str).str.contains(s, na=False))
        ]
        
        if not res.empty:
            st.markdown(f"### Resultados ({len(res)} encontrados)")
            for _, row in res.head(10).iterrows():
                orden = row['ORDEN_LIMP']
                contacto = obtener_contacto(orden)
                
                with st.expander(f"**{orden} - {row['NOMBRE_COMPLETO']}** | {row['AULA']} | DNI: {row['DNI']}", expanded=False):
                    c1, c2 = st.columns(2)
                    
                    with c1:
                        dom = st.text_area("📍 Domicilio", value=contacto.get('domicilio','') if contacto else '', 
                                          key=f"dom_{orden}", placeholder="Calle, número, barrio...")
                        tel_pers = st.text_input("📱 Teléfono Personal", value=contacto.get('telefono_personal','') if contacto else '',
                                                key=f"telp_{orden}", placeholder="2964-XXXXXX")
                        tel_urg = st.text_input("🚨 Teléfono Emergencia", value=contacto.get('telefono_emergencia','') if contacto else '',
                                               key=f"telu_{orden}", placeholder="2964-XXXXXX")
                    
                    with c2:
                        nom_urg = st.text_input("👤 Nombre Contacto Emergencia", 
                                               value=contacto.get('nombre_emergencia','') if contacto else '',
                                               key=f"nomu_{orden}", placeholder="Nombre completo")
                        parent = st.text_input("🔗 Parentesco", value=contacto.get('parentesco_emergencia','') if contacto else '',
                                              key=f"par_{orden}", placeholder="Esposa, Madre, Hermano...")
                        obs = st.text_area("📝 Observaciones", value=contacto.get('observaciones','') if contacto else '',
                                          key=f"obs_{orden}", placeholder="Alergias, grupo sanguíneo, etc.")
                    
                    if st.button("💾 Guardar Contacto", key=f"save_{orden}", type="primary"):
                        guardar_contacto({
                            'orden': int(orden), 'domicilio': dom, 'telefono_personal': tel_pers,
                            'telefono_emergencia': tel_urg, 'nombre_emergencia': nom_urg,
                            'parentesco_emergencia': parent, 'observaciones': obs
                        })
                        st.success(f"✅ Datos guardados para {row['NOMBRE_COMPLETO']}")
                        st.rerun()
    
    st.divider()
    st.subheader("📋 Resumen del Plan de Llamada")
    
    todos = obtener_todos_contactos()
    if todos:
        col1, col2, col3 = st.columns(3)
        with col1: st.metric("Total Registrados", len(todos))
        with col2: 
            con_tel = sum(1 for t in todos if (t.get('telefono_particular') or t.get('telefono_personal')))
            st.metric("Con Teléfono Particular", con_tel)
        with col3:
            con_urg = sum(1 for t in todos if (t.get('telefono_emergencia_licencia') or t.get('telefono_emergencia')))
            st.metric("Con Tel. Emergencia", con_urg)
        
        if st.button("📥 EXPORTAR PLAN DE LLAMADA (EXCEL)", type="primary", use_container_width=True):
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PLAN DE LLAMADA"
            
            headers = [
                "Nro", "APELLIDO Y NOMBRES", "DNI", "CE", "AULA",
                "TRANSP. PUBLICO", "VEHICULO PARTICULAR", "MARCA", "MODELO", "DOMINIO", "TITULAR VEHICULO",
                "LUGAR LICENCIA", "DIRECCION", "BARRIO", "CALLE", "NUMERO", "UNIDAD GN PROXIMA",
                "TEL. PARTICULAR", "TEL. EMERG. LICENCIA", "OBSERV.", "ACTUALIZADO"
            ]
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws['A1'] = "PLAN DE LLAMADA / LICENCIA - ESCUADRÓN H"
            ws['A1'].font = excel_font(bold=True, size=16, color=EXCEL_WHITE)
            ws['A1'].fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
                cell.fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin", color=EXCEL_BORDER), right=Side(style="thin", color=EXCEL_BORDER), top=Side(style="thin", color=EXCEL_BORDER), bottom=Side(style="thin", color=EXCEL_BORDER))
            
            row = 4
            for nro, cont in enumerate(sorted(todos, key=lambda x: x['orden']), 1):
                pers = df[df['ORDEN_LIMP'] == cont['orden']]
                nombre_base = pers.iloc[0]['NOMBRE_COMPLETO'] if not pers.empty else ""
                aula = pers.iloc[0]['AULA'] if not pers.empty else ""
                dni_base = pers.iloc[0]['DNI'] if not pers.empty else ""
                ce_base = pers.iloc[0]['CE'] if not pers.empty else ""
                valores = [
                    nro,
                    cont.get('apellido_nombres') or nombre_base,
                    cont.get('dni') or dni_base,
                    cont.get('ce') or ce_base,
                    aula,
                    cont.get('viaja_transporte_publico',''),
                    cont.get('viaja_vehiculo_particular',''),
                    cont.get('vehiculo_marca',''),
                    cont.get('vehiculo_modelo',''),
                    cont.get('vehiculo_dominio',''),
                    cont.get('vehiculo_titular',''),
                    cont.get('lugar_licencia',''),
                    cont.get('direccion') or cont.get('domicilio',''),
                    cont.get('barrio',''),
                    cont.get('calle',''),
                    cont.get('numero',''),
                    cont.get('unidad_proxima_gn',''),
                    cont.get('telefono_particular') or cont.get('telefono_personal',''),
                    cont.get('telefono_emergencia_licencia') or cont.get('telefono_emergencia',''),
                    cont.get('observaciones',''),
                    cont.get('actualizado_en',''),
                ]
                for c, val in enumerate(valores, 1):
                    ws.cell(row=row, column=c, value=val)
                    ws.cell(row=row, column=c).border = Border(left=Side(style="thin", color=EXCEL_BORDER), right=Side(style="thin", color=EXCEL_BORDER), top=Side(style="thin", color=EXCEL_BORDER), bottom=Side(style="thin", color=EXCEL_BORDER))
                    ws.cell(row=row, column=c).alignment = Alignment(horizontal="center" if c in [1,3,4,5,6,7,16] else "left", vertical="top", wrap_text=True)
                row += 1
            
            widths = [8, 32, 14, 12, 10, 14, 16, 14, 14, 14, 25, 22, 32, 18, 18, 10, 24, 18, 22, 28, 20]
            for idx, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
            
            output = f"PLAN_LLAMADA_{datetime.now().strftime('%d%m%Y')}.xlsx"
            descargar_archivo_auto(excel_bytes(wb), output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success(f"✅ Plan de llamada descargado: **{output}**")
    else:
        st.warning("⚠️ Aún no hay contactos registrados. Usa el buscador para cargar datos.")
  

# --- TAB: RESUMEN ---
with tab_res:
    st.subheader("Resumen General y Novedades")
    mostrar_monitor_novedades()
    st.divider()

    # Recarga de datos
    st.session_state.novedades_lista = obtener_novedades(FECHA_PARTE_STR)
    st.session_state.estado_asistencia = obtener_asistencia(FECHA_PARTE_STR)
    st.session_state.lista_almuerzo = obtener_almuerzo(FECHA_PARTE_STR)

    ambito_resumen = {
        n['orden']: ambito_efectivo(n)
        for n in st.session_state.novedades_lista
    }
    ausentes_resumen = {orden for orden, ambito in ambito_resumen.items() if ambito == "AUSENTE"}
    ausentes_manuales_resumen = {
        orden for orden, estado in st.session_state.estado_asistencia.items()
        if estado == "AUSENTE"
    }
    presentes_manuales_resumen = {
        orden for orden, estado in st.session_state.estado_asistencia.items()
        if estado in {"PRESENTE", "PRESENTE EN INSTITUTO", "PRESENTE EN ESCUADRÓN"}
    }
    total_ausentes_resumen = ausentes_resumen | (ausentes_manuales_resumen - presentes_manuales_resumen)

    # ==========================================================
    # RESUMEN POR AULA
    # ==========================================================

    data_aulas = []

    for aula in AULAS_UNICAS:

        cfg = st.session_state.estado_aulas[aula]
        alumnos = df[df['AULA'] == aula]

        total_aula = len(alumnos)

        ausentes_aula = len(
            {
                row['ORDEN_LIMP']
                for _, row in alumnos.iterrows()
                if row['ORDEN_LIMP'] in total_ausentes_resumen
            }
        )

        presentes_aula = total_aula - ausentes_aula

        almuerzan = sum(
            1
            for _, row in alumnos.iterrows()
            if row['ORDEN_LIMP']
            in st.session_state.lista_almuerzo
        )

        data_aulas.append({
            "Aula": aula,
            "Total": total_aula,
            "Presentes": presentes_aula,
            "Ausentes": ausentes_aula,
            "Almuerzan": almuerzan,
            "Ubicación": (
                cfg.get("ubicacion_m", "-")
                if cfg["estado_m"] == "EN INSTITUTO"
                else "FUERA"
            ),
            "Estado": cfg["estado_m"]
        })

    st.dataframe(
        pd.DataFrame(data_aulas),
        use_container_width=True,
        hide_index=True
    )

    st.divider()

    # =====================================================
    # TABLA PERSONAL AUSENTE O FUERA DE INSTALACIONES
    # =====================================================

    data_ausentes = []

    for nov in st.session_state.novedades_lista:

        alumno_df = df[df['ORDEN_LIMP'] == nov['orden']]

        if not alumno_df.empty:

            alumno = alumno_df.iloc[0]

            data_ausentes.append({
                "Nombre": alumno["NOMBRE_COMPLETO"],
                "Motivo": nov["estado"],
                "Desde": nov["fecha_ini"],
                "Hasta": nov["fecha_fin"],
            })

    # AUSENTES MANUALES
    for orden, estado in st.session_state.estado_asistencia.items():

        if estado == "AUSENTE":

            # Evitar duplicados si ya tiene novedad cargada
            ya_existe = any(
                nov["orden"] == orden
                for nov in st.session_state.novedades_lista
            )

            if not ya_existe:

                alumno_df = df[df['ORDEN_LIMP'] == orden]

                if not alumno_df.empty:

                    alumno = alumno_df.iloc[0]

                    data_ausentes.append({
                        "Nombre": alumno["NOMBRE_COMPLETO"],
                        "Motivo": "AUSENTE",
                        "Desde": FECHA_PARTE_STR,
                        "Hasta": FECHA_PARTE_STR
                    })

    df_ausentes = pd.DataFrame(data_ausentes)

    if not df_ausentes.empty:
        df_ausentes = df_ausentes.sort_values(["Motivo", "Nombre"]).reset_index(drop=True)
        df_ausentes.insert(0, "Nro", range(1, len(df_ausentes) + 1))
        st.dataframe(
            df_ausentes,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.success("Sin personal ausente.")

with tab_res:
    st.divider()
    st.subheader("Minuta informativa")
    formato_minuta = st.radio(
        "Formato de minuta:",
        ["Visual WhatsApp / Celular", "Formal cl\u00e1sico"],
        index=0,
        horizontal=True,
        key="formato_minuta",
    )
    minuta_texto = generar_minuta_informativa(formato_minuta)
    st.text_area("Minuta generada autom?ticamente desde Novedades", value=minuta_texto, height=560)
    nombre_minuta = f"MINUTA_ESCUADRON_H_{st.session_state.fecha_reporte.strftime('%d%m%Y')}.txt"
    minuta_bytes = minuta_texto.encode("utf-8-sig")
    st.download_button(
        "Descargar minuta (.txt)",
        data=minuta_bytes,
        file_name=nombre_minuta,
        mime="text/plain; charset=utf-8",
        key=f"descargar_minuta_{st.session_state.fecha_reporte.strftime('%Y%m%d')}_{len(minuta_bytes)}",
        on_click="ignore",
        use_container_width=True
    )


    st.divider()
    st.subheader("Historial de movimientos")
    col_fi, col_ff = st.columns(2)
    with col_fi:
        fecha_desde_hist = st.date_input("Desde", datetime.now().date(), key="fecha_historial_desde")
    with col_ff:
        fecha_hasta_hist = st.date_input("Hasta", datetime.now().date(), key="fecha_historial_hasta")

    movimientos = obtener_movimientos()
    if movimientos:
        df_mov = pd.DataFrame(movimientos)
        df_mov = df_mov[df_mov.get("modulo", "") == "Novedades"].copy()
        df_mov = df_mov.rename(columns={
            "fecha_hora": "Fecha/hora",
            "fecha_parte": "Fecha parte",
            "modulo": "Módulo",
            "accion": "Acción",
            "orden": "Orden interno",
            "nombre": "Nombre completo",
            "aula": "Aula",
            "detalle": "Detalle",
        })
        detalle_cols = df_mov["Detalle"].apply(parsear_detalle_movimiento).apply(pd.Series)
        df_mov = pd.concat([df_mov, detalle_cols], axis=1)
        nombres_separados = df_mov["Nombre completo"].fillna("").astype(str).str.strip().str.split(n=1, expand=True)
        df_mov["Apellido"] = nombres_separados[0].fillna("").astype(str).str.strip().str.rstrip(",")
        df_mov["Nombre"] = nombres_separados[1].fillna("") if nombres_separados.shape[1] > 1 else ""
        df_mov["Nombre"] = df_mov["Nombre"].astype(str).str.strip().str.lstrip(",").str.strip()
        datos_personal = df[["ORDEN_LIMP", "DNI", "CE"]].copy()
        df_mov["Orden personal"] = pd.to_numeric(df_mov["Orden interno"], errors="coerce")
        df_mov = df_mov.merge(datos_personal, left_on="Orden personal", right_on="ORDEN_LIMP", how="left")
        df_mov["DNI"] = df_mov["DNI"].fillna("").astype(str).str.replace(".0", "", regex=False)
        df_mov["CE"] = df_mov["CE"].fillna("").astype(str).str.replace(".0", "", regex=False)
        df_mov["Fecha parte dt"] = pd.to_datetime(df_mov["Fecha parte"], errors="coerce").dt.date
        df_mov = df_mov[
            (df_mov["Fecha parte dt"] >= fecha_desde_hist) &
            (df_mov["Fecha parte dt"] <= fecha_hasta_hist)
        ].copy()
        df_mov["Clave aspirante"] = df_mov["Orden interno"].fillna(df_mov["Nombre completo"]).astype(str)
        df_mov = df_mov.sort_values("Fecha/hora", ascending=False).copy()

        if not df_mov.empty:
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                motivos_sel = st.multiselect("Motivo", sorted([m for m in df_mov["Motivo"].dropna().unique() if m]), placeholder="Filtrar")
            with fc2:
                presencias_sel = st.multiselect("Presencia", sorted([p for p in df_mov["Presencia"].dropna().unique() if p]), placeholder="Filtrar")
            with fc3:
                aulas_sel = st.multiselect("Aula", sorted([a for a in df_mov["Aula"].dropna().unique() if a]), placeholder="Filtrar")

            if motivos_sel:
                df_mov = df_mov[df_mov["Motivo"].isin(motivos_sel)]
            if presencias_sel:
                df_mov = df_mov[df_mov["Presencia"].isin(presencias_sel)]
            if aulas_sel:
                df_mov = df_mov[df_mov["Aula"].isin(aulas_sel)]

            if not df_mov.empty:
                opciones_asp = df_mov[["Clave aspirante", "Nombre completo", "Aula"]].drop_duplicates("Clave aspirante")
                opciones_asp = opciones_asp.sort_values("Nombre completo")
                etiquetas_asp = {
                    f"{row['Nombre completo']} | Aula: {row['Aula']}": row["Clave aspirante"]
                    for _, row in opciones_asp.iterrows()
                }
                aspirante_sel = st.selectbox(
                    "Buscar aspirante",
                    list(etiquetas_asp.keys()),
                    index=None,
                    placeholder="Escribi para buscar o deja vacio para ver todo",
                    key="historial_aspirante_sel",
                    help="Si no seleccionas un aspirante, se muestra todo el listado del rango de fechas"
                )
                if aspirante_sel:
                    df_mov = df_mov[df_mov["Clave aspirante"] == etiquetas_asp[aspirante_sel]]

        columnas = ["Fecha/hora", "Apellido", "Nombre", "DNI", "CE", "Aula", "Motivo", "Presencia", "Desde", "Hasta", "Observación"]

        if not df_mov.empty:
            st.caption(f"{len(df_mov)} movimiento(s) encontrados")
            st.dataframe(df_mov[columnas], use_container_width=True, hide_index=True)

            edit_mov_id = st.session_state.get("editando_movimiento_id")
            if edit_mov_id:
                mov_edit = df_mov[df_mov["id"] == edit_mov_id]
                if not mov_edit.empty:
                    mov = mov_edit.iloc[0]
                    with st.form("form_editar_movimiento"):
                        st.markdown(f"### Editar movimiento de {mov.get('Nombre completo', '-')}")
                        e1, e2, e3 = st.columns(3)
                        with e1:
                            mov_motivo = st.text_input("Motivo", value=str(mov.get("Motivo", "")), key="mov_edit_motivo")
                        with e2:
                            opciones_presencia = ["Ausente", "Presente en instituto", "Presente en escuadrón"]
                            presencia_actual = str(mov.get("Presencia", "Ausente"))
                            indice_presencia = opciones_presencia.index(presencia_actual) if presencia_actual in opciones_presencia else 0
                            mov_presencia = st.selectbox("Presencia", opciones_presencia, index=indice_presencia, key="mov_edit_presencia")
                        with e3:
                            mov_aula = st.text_input("Aula", value=str(mov.get("Aula", "")), key="mov_edit_aula")
                        e4, e5 = st.columns(2)
                        with e4:
                            mov_desde = st.text_input("Desde", value=str(mov.get("Desde", "")), key="mov_edit_desde")
                        with e5:
                            mov_hasta = st.text_input("Hasta", value=str(mov.get("Hasta", "")), key="mov_edit_hasta")
                        mov_obs = st.text_input("Observación", value=str(mov.get("Observación", "")), key="mov_edit_obs")
                        guardar_mov, cancelar_mov = st.columns([3, 1])
                        with guardar_mov:
                            guardar = st.form_submit_button("Guardar movimiento", type="primary", use_container_width=True)
                        with cancelar_mov:
                            cancelar = st.form_submit_button("Cancelar", use_container_width=True)
                        if guardar:
                            detalle_nuevo = f"{mov_motivo.upper()} | {mov_presencia} | {mov_desde.upper()} a {mov_hasta.upper()} | {mov_obs.upper()}"
                            orden_valor = mov.get("Orden interno")
                            try:
                                orden_valor = int(orden_valor) if pd.notna(orden_valor) and str(orden_valor).strip() else None
                            except Exception:
                                orden_valor = None
                            actualizar_movimiento(int(mov["id"]), {
                                "fecha_parte": mov.get("Fecha parte"),
                                "modulo": mov.get("Módulo", "Historial"),
                                "accion": "EDITAR MOVIMIENTO",
                                "orden": orden_valor,
                                "nombre": mov.get("Nombre completo"),
                                "aula": mov_aula.upper(),
                                "detalle": detalle_nuevo,
                            })
                            st.session_state.editando_movimiento_id = None
                            st.success("Movimiento actualizado")
                            st.rerun()
                        if cancelar:
                            st.session_state.editando_movimiento_id = None
                            st.rerun()

            with st.expander("Administrar movimientos filtrados", expanded=False):
                st.caption("Desde acá podés editar o borrar registros de prueba del historial.")
                for _, mov in df_mov.iterrows():
                    mov_id = int(mov["id"])
                    c_info, c_edit, c_del = st.columns([6, 1, 1])
                    with c_info:
                        st.markdown(f"**{mov.get('Apellido', '')} {mov.get('Nombre', '')}** | {mov.get('Motivo', '-')} | {mov.get('Presencia', '-')}")
                        st.caption(f"{mov.get('Fecha/hora', '-')} | Aula {mov.get('Aula', '-')} | {mov.get('Desde', '-')} a {mov.get('Hasta', '-')} | {mov.get('Observación', '')}")
                    with c_edit:
                        if st.button("Editar", key=f"edit_mov_{mov_id}", use_container_width=True):
                            st.session_state.editando_movimiento_id = mov_id
                            st.rerun()
                    with c_del:
                        if st.button("Eliminar", key=f"del_mov_{mov_id}", use_container_width=True):
                            eliminar_movimiento(mov_id)
                            if st.session_state.get("editando_movimiento_id") == mov_id:
                                st.session_state.editando_movimiento_id = None
                            st.toast("Movimiento eliminado")
                            st.rerun()
        else:
            st.info("No hay movimientos para los filtros seleccionados.")

        historial_df = df_mov[columnas] if not df_mov.empty else pd.DataFrame(columns=columnas)
        wb_hist = openpyxl.Workbook()
        ws_hist = wb_hist.active
        ws_hist.title = "HISTORIAL"
        from openpyxl.styles import Alignment, PatternFill, Border, Side

        titulo_fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
        subtitulo_fill = PatternFill(start_color=EXCEL_OLIVE_LIGHT, end_color=EXCEL_OLIVE_LIGHT, fill_type="solid")
        header_fill = PatternFill(start_color=EXCEL_OLIVE, end_color=EXCEL_OLIVE, fill_type="solid")
        thin = Side(style="thin", color=EXCEL_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws_hist.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        ws_hist.cell(1, 1, "ESCUADRON H \"Cabo Marcelo Godoy\"")
        ws_hist.cell(1, 1).font = excel_font(bold=True, size=16, color=EXCEL_WHITE)
        ws_hist.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_hist.cell(1, 1).fill = titulo_fill
        ws_hist.row_dimensions[1].height = 28

        ws_hist.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
        ws_hist.cell(2, 1, "Historial De Movimientos de Aspirantes")
        ws_hist.cell(2, 1).font = excel_font(bold=True, size=13, color=EXCEL_OLIVE_DARK)
        ws_hist.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_hist.cell(2, 1).fill = subtitulo_fill
        ws_hist.row_dimensions[2].height = 24

        ws_hist.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columnas))
        ws_hist.cell(3, 1, f"Periodo: {fecha_desde_hist.strftime('%d/%m/%Y')} al {fecha_hasta_hist.strftime('%d/%m/%Y')} | Registros: {len(historial_df)}")
        ws_hist.cell(3, 1).font = excel_font(italic=True, size=10, color=EXCEL_TEXT_MUTED)
        ws_hist.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center")

        header_row = 5
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws_hist.cell(header_row, col_idx, col_name)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx, (_, row) in enumerate(historial_df.iterrows(), header_row + 1):
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws_hist.cell(row_idx, col_idx, row.get(col_name, ""))
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=EXCEL_OLIVE_ROW, end_color=EXCEL_OLIVE_ROW, fill_type="solid")

        widths = {
            "Fecha/hora": 20, "Apellido": 18, "Nombre": 28, "DNI": 14, "CE": 12,
            "Aula": 12, "Motivo": 18, "Presencia": 22, "Desde": 14, "Hasta": 14, "Observación": 36
        }
        for col_idx, col_name in enumerate(columnas, 1):
            ws_hist.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(col_name, 16)
        ws_hist.freeze_panes = "A6"
        ws_hist.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(len(columnas))}{max(header_row, header_row + len(historial_df))}"

        st.download_button(
            "descargar historial de movimiento",
            data=excel_bytes(wb_hist),
            file_name=f"HISTORIAL_MOVIMIENTOS_{fecha_desde_hist.strftime('%d%m%Y')}_{fecha_hasta_hist.strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("No hay movimientos registrados.")

with tab_res:
    if True:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PARTE DIARIO"
    
        thin = Side(style="thin", color=EXCEL_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
        sub_fill = PatternFill(start_color=EXCEL_OLIVE_LIGHT, end_color=EXCEL_OLIVE_LIGHT, fill_type="solid")
        fecha_titulo = st.session_state.fecha_reporte.strftime('%d%b%y').upper()
        dia_reporte = DIAS_SEMANA[st.session_state.fecha_reporte.weekday()]
    
        ws.merge_cells('A1:J1')
        ws['A1'] = f"PARTE DIARIO DEL ESCUADRÓN H - {fecha_titulo}"
        ws['A1'].font = excel_font(bold=True, size=16, color=EXCEL_WHITE)
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
    
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Día: {dia_reporte} | Primera obligación: 06:00 hs | Generado: {datetime.now().strftime('%H:%M')}"
        ws['A2'].font = excel_font(italic=True, size=11, color=EXCEL_TEXT_MUTED)
        ws['A2'].alignment = Alignment(horizontal="center")
    
        metric_headers = [
            "TOTAL", "EN INSTITUTO", "EN ESCUADRÓN", "AUSENTES",
            "GUARDIA D.", "GUARDIA N.", "COMISIÓN",
            "1RA OBLIG. 06:00", "3ER AÑO", "AOP"
        ]
        metric_values = [
            TOTAL_ESCUADRON, en_instituto, presentes_escuadron, len(total_ausentes),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'ENTRANTE GUARDIA DIURNA'),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'ENTRANTE GUARDIA NOCTURNA'),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'COMISIÓN'),
            primera_total, primera_tercer_anio, primera_aop
        ]
        for col, label in enumerate(metric_headers, 1):
            cell = ws.cell(row=4, column=col, value=label)
            cell.font = excel_font(bold=True, size=9, color=EXCEL_TEXT_DARK)
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            val = ws.cell(row=5, column=col, value=metric_values[col - 1])
            val.font = excel_font(bold=True, size=12)
            val.alignment = Alignment(horizontal="center")
            val.border = border
    
        ws.merge_cells('A8:J8')
        ws['A8'] = "NOVEDADES DEL PERSONAL (AUSENTES JUSTIFICADOS)"
        ws['A8'].font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
    
        nov_headers = ["Nro", "GRADO", "APELLIDO Y NOMBRE", "DNI", "CE", "NOVEDAD", "DETALLE", "DESDE", "HASTA", "AULA"]
        for col, h in enumerate(nov_headers, 1):
            cell = ws.cell(row=9, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
    
        current_row = 10
        if st.session_state.novedades_lista:
            for i, nov in enumerate(st.session_state.novedades_lista, 1):
                row = current_row + i - 1
                values = [
                    i, nov.get('grado', ''), nov['nombre'], nov.get('dni', ''), nov.get('ce', ''),
                    nov['estado'], nov['detalle'], nov['fecha_ini'], nov['fecha_fin'], nov.get('aula', '-')
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center" if col in [1, 2, 4, 5, 6, 8, 9, 10] else "left")
            current_row += len(st.session_state.novedades_lista)
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1, value="Sin novedades registradas en la guardia").font = excel_font(italic=True, color=EXCEL_TEXT_MUTED)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
            current_row += 1
    
        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value=f"HORARIOS DE INGRESO - {dia_reporte.upper()}").font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
        current_row += 1
    
        aulas_0600 = []
        for aula in AULAS_UNICAS:
            hor = st.session_state.horarios_config[aula]
            if normalizar_hora_ingreso(hor.get('ent_m')) == '06:00':
                aulas_0600.append(aula)
    
        texto_0600 = (
            f"• Ingreso 06:00 hs (Primera obligación): Aula(s) {', '.join(aulas_0600) if aulas_0600 else 'N/A'} "
            f"— Forman {primera_total} aspirante(s): {primera_tercer_anio} de 3er año y {primera_aop} de AOP."
        )
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value=texto_0600)
        ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
        current_row += 1
    
        otros_ingresos = []
        for aula in AULAS_UNICAS:
            hor = st.session_state.horarios_config[aula]
            if normalizar_hora_ingreso(hor.get('ent_m')) != '06:00':
                cant = len(df_presentes_escuadron_base[df_presentes_escuadron_base['AULA'] == aula])
                otros_ingresos.append(f"{aula}: {hor.get('ent_m')} hs ({cant})")
        if otros_ingresos:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1, value=f"• Otros ingresos mañana: {', '.join(otros_ingresos)}")
            ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 1
    
        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value="OBSERVACIONES").font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
    
        for col, width in enumerate([10, 14, 34, 12, 10, 18, 24, 12, 12, 12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
        output = f"PARTE_DIARIO_ESCUADRON_H_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        st.download_button(
            "GENERAR PARTE DIARIO (EXCEL)",
            data=excel_bytes(wb),
            file_name=output,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            key="download_parte_diario_formal"
        )
    
    # ==============================================================================
    # 5. EXPORTAR EXCEL (PARTE DIARIO DETALLADO)
    # ==============================================================================
    if True:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PARTE DIARIO DETALLADO"

        thin = Side(style="thin", color=EXCEL_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color=EXCEL_OLIVE_DARK, end_color=EXCEL_OLIVE_DARK, fill_type="solid")
        sub_fill = PatternFill(start_color=EXCEL_OLIVE_LIGHT, end_color=EXCEL_OLIVE_LIGHT, fill_type="solid")
        fecha_titulo = st.session_state.fecha_reporte.strftime('%d%b%y').upper()
        dia_reporte = DIAS_SEMANA[st.session_state.fecha_reporte.weekday()]

        ws.merge_cells('A1:J1')
        ws['A1'] = f"PARTE DIARIO DETALLADO DEL ESCUADRÓN H - {fecha_titulo}"
        ws['A1'].font = excel_font(bold=True, size=16, color=EXCEL_WHITE)
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells('A2:J2')
        ws['A2'] = f"Día: {dia_reporte} | Primera obligación: 06:00 hs | Generado: {datetime.now().strftime('%H:%M')}"
        ws['A2'].font = excel_font(italic=True, size=11, color=EXCEL_TEXT_MUTED)
        ws['A2'].alignment = Alignment(horizontal="center")

        metric_headers = [
            "TOTAL", "EN INSTITUTO", "EN ESCUADRÓN", "AUSENTES",
            "GUARDIA D.", "GUARDIA N.", "COMISIÓN",
            "1RA OBLIG. 06:00", "3ER AÑO", "AOP"
        ]
        metric_values = [
            TOTAL_ESCUADRON, en_instituto, presentes_escuadron, len(total_ausentes),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'ENTRANTE GUARDIA DIURNA'),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'ENTRANTE GUARDIA NOCTURNA'),
            sum(1 for n in st.session_state.novedades_lista if n['estado'] == 'COMISIÓN'),
            primera_total, primera_tercer_anio, primera_aop
        ]
        for col, label in enumerate(metric_headers, 1):
            cell = ws.cell(row=4, column=col, value=label)
            cell.font = excel_font(bold=True, size=9, color=EXCEL_TEXT_DARK)
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            val = ws.cell(row=5, column=col, value=metric_values[col - 1])
            val.font = excel_font(bold=True, size=12)
            val.alignment = Alignment(horizontal="center")
            val.border = border

        ws.merge_cells('A8:J8')
        ws['A8'] = "NOVEDADES DEL PERSONAL"
        ws['A8'].font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)

        nov_headers = ["Nro", "GRADO", "APELLIDO Y NOMBRE", "DNI", "CE", "NOVEDAD", "PRESENCIA", "DETALLE", "DESDE/HASTA", "AULA"]
        for col, h in enumerate(nov_headers, 1):
            cell = ws.cell(row=9, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        current_row = 10
        if st.session_state.novedades_lista:
            for i, nov in enumerate(st.session_state.novedades_lista, 1):
                row = current_row + i - 1
                values = [
                    i, nov.get('grado', ''), nov['nombre'], nov.get('dni', ''), nov.get('ce', ''),
                    nov['estado'], AMBITOS_NOVEDAD.get(ambito_efectivo(nov), ""), nov['detalle'],
                    f"{nov['fecha_ini']} a {nov['fecha_fin']}", nov.get('aula', '-')
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center" if col in [1, 2, 4, 5, 6, 7, 9, 10] else "left")
            current_row += len(st.session_state.novedades_lista)
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1, value="Sin novedades registradas en la guardia").font = excel_font(italic=True, color=EXCEL_TEXT_MUTED)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
            current_row += 1

        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value="PERSONAL QUE ALMUERZA").font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
        current_row += 1
        alm_headers = ["Nro", "APELLIDO Y NOMBRE", "AULA", "CE", "DNI"]
        for col, h in enumerate(alm_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        current_row += 1
        if st.session_state.lista_almuerzo:
            for i, orden in enumerate(sorted(st.session_state.lista_almuerzo), 1):
                asp = df[df['ORDEN_LIMP'] == orden]
                if not asp.empty:
                    p = asp.iloc[0]
                    values = [i, p['NOMBRE_COMPLETO'], p['AULA'], p['CE'], p['DNI']]
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
                    current_row += 1
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1, value="Sin personal cargado para almuerzo").font = excel_font(italic=True, color=EXCEL_TEXT_MUTED)
            current_row += 1

        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value=f"HORARIOS DE INGRESO - {dia_reporte.upper()}").font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
        current_row += 1
        hor_headers = ["AULA", "ENT. MAÑANA", "SAL. MAÑANA", "ENT. TARDE", "SAL. TARDE"]
        for col, h in enumerate(hor_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        current_row += 1
        for aula in AULAS_UNICAS:
            hor = st.session_state.horarios_config[aula]
            values = [aula, hor.get('ent_m'), hor.get('sal_m'), hor.get('ent_t'), hor.get('sal_t')]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value="CONTROL DE AULAS").font = excel_font(bold=True, size=12, color=EXCEL_OLIVE_DARK)
        current_row += 1
        aula_headers = ["AULA", "TOTAL", "PRESENTES", "AUSENTES", "ALMUERZAN", "UBICACIÓN", "ESTADO M", "ESTADO T"]
        for col, h in enumerate(aula_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = excel_font(bold=True, color=EXCEL_WHITE, size=9)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        current_row += 1
        for aula in AULAS_UNICAS:
            cfg = st.session_state.estado_aulas[aula]
            alumnos = df[df['AULA'] == aula]
            ausentes_aula = len({row['ORDEN_LIMP'] for _, row in alumnos.iterrows() if row['ORDEN_LIMP'] in total_ausentes})
            almuerzan = sum(1 for _, row in alumnos.iterrows() if row['ORDEN_LIMP'] in st.session_state.lista_almuerzo)
            values = [
                aula, len(alumnos), len(alumnos) - ausentes_aula, ausentes_aula, almuerzan,
                cfg.get("ubicacion_m", "-") if cfg["estado_m"] == "EN INSTITUTO" else "FUERA",
                cfg["estado_m"], cfg["estado_t"]
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

        for col, width in enumerate([10, 16, 34, 12, 10, 18, 18, 24, 18, 12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        output = f"PARTE_DIARIO_DETALLADO_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        st.download_button(
            "GENERAR PARTE DIARIO DETALLADO (EXCEL)",
            data=excel_bytes(wb),
            file_name=output,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
            key="download_parte_diario_detallado"
        )
